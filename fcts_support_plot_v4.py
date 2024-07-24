from copy import deepcopy

import cartopy.crs as ccrs
import matplotlib.colors as plcol
import matplotlib.pyplot as plt
import numpy as np
import regionmask
import scipy.spatial as sspa
import seaborn as sns
import xarray as xr
from matplotlib import cm
from matplotlib.cm import ScalarMappable
from matplotlib.patches import (ArrowStyle, FancyArrowPatch, FancyBboxPatch,
                                Polygon)
from mpl_toolkits.axes_grid1.inset_locator import inset_axes

CB_color_cycle = sns.color_palette( 'colorblind', n_colors=10000 )
import openpyxl
import pandas as pd

from fcts_support_basic import *
from fcts_support_event import *
from fcts_support_synthesis import *
from fcts_support_training import *

dict_shorten_entities = {}

dict_even_shorter_entities = {'China, Peoples Rep. (coal & cement)':'China (coal & cement)', 'Royal Dutch Shell, The Netherlands':'Shell'}


# order based on cumulative CO2 emissions over 1850-2020
list_color_entities = ['Czechoslovakia', 'APA Corporation', 'Whitehaven Coal',
                       'Pioneer Natural Resources', 'PTTEP', 'BASF', 'Saudi Aramco', 'Cemex',
                       'Adani Enterprises', 'CNPC', 'Kiewit Mining Group', 'Lukoil', 'Chevron',
                       'Abu Dhabi National Oil Company', 'Pertamina', 'North Korea',
                       'EOG Resources', 'Kazakhstan', 'TurkmenGaz', 'Surgutneftegas',
                       'Taiheiyo Cement', 'Bapco Energies', 'Ecopetrol', 'Inpex',
                       'Devon Energy', 'Cenovus Energy', 'Alliance Resource Partners',
                       'Cyprus AMAX Minerals', 'Equinor', 'Vale', 'UK Coal', 'CONSOL Energy',
                       'EQT Corporation', 'Anglo American', 'OMV Group', 'RWE',
                       'Exxaro Resources Ltd', 'Occidental Petroleum', 'Novatek',
                       'ConocoPhillips', 'Westmoreland Mining', 'Arch Resources', 'Tullow Oil',
                       'Repsol', 'Ovintiv', 'China (Coal)', 'ExxonMobil', 'Wolverine Fuels',
                       'Bumi Resources', 'Petroleos de Venezuela', 'Kuwait Petroleum Corp.',
                       'Marathon Oil', 'Cloud Peak', 'Libya National Oil Corp.', 'Petrobras',
                       'Coal India', 'Canadian Natural Resources', 'Coterra Energy',
                       'Heidelberg Materials', 'Tourmaline Oil', 'QatarEnergy',
                       'China (Cement)', 'Eni', 'Petroleum Development Oman',
                       'Navajo Transitional Energy Company', 'Woodside Energy',
                       'Nigerian National Petroleum Corp.', 'Petronas', 'Slovakia',
                       'Murphy Oil', 'Sinopec', 'Syrian Petroleum', 'CNOOC',
                       'Russian Federation', 'Peabody Coal Group', 'Former Soviet Union',
                       'Shell', 'Pemex', 'Glencore', 'American Consolidated Natural Resources',
                       'TotalEnergies', 'YPF', 'Ukraine', 'Iraq National Oil Company',
                       'British Coal Corporation', 'Egyptian General Petroleum',
                       'Singareni Collieries', 'Sonangol', 'Orlen', 'Rio Tinto', 'Gazprom',
                       'Continental Resources', 'BHP', 'North American Coal', 'CNX Resources',
                       'CRH', 'Santos', 'PetroEcuador', 'Adaro Energy', 'Vistra', 'ONGC India',
                       'SM Energy', 'Southwestern Energy', 'Sonatrach',
                       'National Iranian Oil Co.', 'Banpu', 'Seriti Resources',
                       'Suncor Energy', 'Antero', 'Obsidian Energy', 'Holcim Group',
                       'Czech Republic', 'Petoro', 'Sasol', 'BP', 'Chesapeake Energy',
                       'Rosneft', 'Poland', 'Teck Resources', 'Hess Corporation',
                       'Alpha Metallurgical Resources', 'Naftogaz'
                      ]
dico_color_entities = {ent:CB_color_cycle[i_ent] for i_ent, ent in enumerate(list_color_entities)}


list_letters_panels = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']


dico_add_emdatcountries_regionmask = {'Bosnia and Herzegovina':['Bosnia and Herz.'],\
                                      'Bolivia (Plurinational State of)':['Bolivia'],\
                                      'Canary Is':['Spain'],\
                                      'Canary Islands':['Spain'],\
                                      'Czech Republic (the)':['Czechia'],\
                                      "Democratic People's Republic of Korea":['North Korea'],\
                                      "Korea (the Democratic People's Republic of)":['North Korea'],\
                                      'Korea (the Republic of)':['South Korea'],\
                                      'Malta':['Cyprus'],\
                                      'Macedonia (the former Yugoslav Republic of)':['North Macedonia'],\
                                      'Netherlands (the)':['Netherlands'],\
                                      'Netherlands (Kingdom of the)':['Netherlands'],\
                                      'Republic of Korea':['South Korea'],\
                                      'Russian Federation (the)':['Russia'],\
                                      'Russian Federation':['Russia'],\
                                      'Serbia Montenegro':['Serbia', 'Montenegro'],\
                                      'Sudan (the)':['Sudan'],\
                                      'Türkiye':['Turkey'],\
                                      'United Kingdom of Great Britain and Northern Ireland (the)':['United Kingdom'],\
                                      'United Kingdom of Great Britain and Northern Ireland':['United Kingdom'],\
                                      'United States of America (the)':['United States of America'] }









#---------------------------------------------------------------------------------------------------------------------------
# FIGURE 1
#---------------------------------------------------------------------------------------------------------------------------
class figure1_v4:
    #--------------------------------------------------------------
    # BASIC FUNCTIONS
    def __init__( self, results, ind_evts, emdat, data_to_do, name_data, width_figure=21 ):
        self.results = results
        self.ind_evts = ind_evts
        self.emdat = emdat
        self.data_to_do = data_to_do
        self.name_data = name_data
        self.width_figure = width_figure
        self.fontsize = 20
        self.fontsizes = {'global_cbar_label':self.fontsize, 'global_cbar_ticks':0.8*self.fontsize, 'letter_panel':0.9*self.fontsize,\
                          'reg_cbar_label':0.7*self.fontsize, 'reg_cbar_ticks':0.7*self.fontsize,\
                          'gmt_legend':0.65*self.fontsize, 'gmt_legend_title':0.75*self.fontsize, 'gmt_text':0.70*self.fontsize, 'gmt_axis_ticks':0.65*self.fontsize,\
                          'gmt_axis_ticks_bars':0.7*self.fontsize, 'gmt_title':0.675*self.fontsize, 'gmt_axis_labels':0.70*self.fontsize, \
                          'attrib_legend':0.70*self.fontsize, 'attrib_legend_title':0.75*self.fontsize, 'attrib_axis_ticks':0.7*self.fontsize, 'attrib_axis_labels':0.8*self.fontsize, 'attrib_txt_period':0.6*self.fontsize }
        self.dico_lw = {'rect_cases':3, 'global_map':3, 'arrow':2, \
                        'reg_map_country':2, 'reg_map_reg':3, \
                        'gmt_central':4, 'gmt_uncertainties':1, 'gmt_event':3, 'gmt_median_ip':5,\
                        'return_wwo':3, 'return_lines':3, 'return_arrows':3}
        self.arrow_style = ArrowStyle('Fancy', head_length=0.5, head_width=0.5, tail_width=0.25)#ArrowStyle('Fancy', head_length=0.5, head_width=0.5, tail_width=0.25)
        self.dico_cols = {'with_CC':CB_color_cycle[0], 'without_CC':CB_color_cycle[8], 'event':'k', 'counter_event':'grey', 'attrib':'r', 'attrib_median':'darkred', 'uncertainties':'grey', 'reg_map_country':'k', 'reg_map_event':'fuchsia', 'letter_panel_face':(0.95,0.95,0.95,0.5), 'letter_panel_edge':(0.5,0.5,0.5)}
        self.width_ratios = [0.60, 1,2, 0.75, 1,2, 0.125] # [1.10, 1,2, 0.75, 1,2, 0.125]
        self.height_ratios = [1, 3, 1]
        self.shift_subplot_bottom, self.shift_subplot_upper = 0.055, -0.020#0.030, -0.010 (0.45??)
        self.shift_subplot_reg = 0.01
        self.factor_sizes_reg = 1.10#22#39
        self.shift_axbars_gmt_int, self.shift_axbars_gmt_prb, self.width_axbars_gmt = 0.0075, 0.010, 0.0125
        
    def plot( self, path_save, attrib_or_gmt='gmt', arrow_or_label='both' ):
        self.fig = plt.figure( figsize=(self.width_figure, self.width_figure) )
        self.spec = gridspec.GridSpec(ncols=7, nrows=3, figure=self.fig, width_ratios=self.width_ratios, height_ratios=self.height_ratios, left=0.025, right=0.975, bottom=0.035, top=0.970, wspace=0.25, hspace=0.05 )

        # Plot central map
        ax_map = plt.subplot( self.spec[1,:-1], projection=ccrs.Robinson(central_longitude=0) )
        ax_map.set_global() # necessary for FancyArrowPatch
        self.plot_map_country_events( ax=ax_map )#, cbaxescbaxes_map )
        t = ax_map.text(x=0.025, y=0.50, s='('+list_letters_panels[0]+')', transform=ax_map.transAxes, fontdict={'size':self.fontsizes['letter_panel']}, weight='bold',\
                        verticalalignment='center', horizontalalignment='center', rotation_mode='anchor',\
                        bbox={'boxstyle':'round', 'facecolor':self.dico_cols['letter_panel_face'], 'edgecolor':self.dico_cols['letter_panel_edge'], 'alpha':1})

        # Preparation for coords of subaxes
        self.ind_row_bottom, self.ind_row_upper = 2, 0
        self.dico_subplot_evts = {'2021-0390-USA':np.array([self.ind_row_upper, 1]), '2003-0391-FRA':np.array([self.ind_row_upper, 4]),\
                                  '2010-0206-IND':np.array([self.ind_row_bottom, 1]), '2022-0248-IND':np.array([self.ind_row_bottom, 1]), '2013-0582-CHN':np.array([self.ind_row_bottom, 4])}
        self.dico_shifts_subplots = {self.ind_row_bottom:self.shift_subplot_bottom, self.ind_row_upper:self.shift_subplot_upper}

        # looping on events to plot
        for ind_evt in self.ind_evts:
            # preparing identification of panel
            self.letter_panel = list_letters_panels[{'2021-0390-USA':1, '2003-0391-FRA':2, '2010-0206-IND':3, '2022-0248-IND':3, '2013-0582-CHN':4}[ind_evt]]
            
            # preparing event
            self.evt_id, self.evt_fits = deepcopy(self.results[ind_evt][0]), deepcopy(self.results[ind_evt][3]) # [ evt_id, datasets_evt, evt_obs, evt_fits, synth ]
            obs_selecttime = self.evt_id.select_time( self.data_to_do[self.name_data][0] )
            
            # checking that only one window, otherwise will cause issues
            if len(self.evt_id.windows) > 1:
                raise Exception("For this figure, supposed to have 1 window")
            else:
                self.window = self.evt_id.windows[0]
                
            # preparing coordinates
            self.country, self.mask, self.central_longitude, self.lon_borders, self.lat_borders = self.id_map_event(obs_selecttime=obs_selecttime, type_reg='country', set_ratio_lon_lat=1)
            _, _, self.central_longitude_evt, self.lon_borders_evt, self.lat_borders_evt = self.id_map_event(obs_selecttime=obs_selecttime, type_reg='event')
            
            # reducing size of window to avoid small region within big country:
            margin_lat_map, margin_lon_map, margin_lat_mask, margin_lon_mask = 5, 10, 1, 2
            self.lat_borders[0] = np.max( [self.lat_borders[0], self.lat_borders_evt[0]-margin_lat_map] )
            self.lat_borders[1] = np.min( [self.lat_borders[1], self.lat_borders_evt[1]+margin_lat_map] )
            self.lon_borders[0] = np.max( [self.lon_borders[0], self.lon_borders_evt[0]-margin_lon_map] )
            self.lon_borders[1] = np.min( [self.lon_borders[1], self.lon_borders_evt[1]+margin_lon_map] )
            self.mask = self.mask * xr.where((self.mask.lat>=self.lat_borders[0]-margin_lat_mask) & (self.mask.lat<=self.lat_borders[1]+margin_lat_mask) &\
                                             (self.mask.lon>=self.lon_borders[0]-margin_lon_mask) & (self.mask.lon<=self.lon_borders[1]+margin_lon_mask), 1, 0)

            # plot return periods
            if attrib_or_gmt == 'attrib':
                subax_attribgmt, bbox_2 = self.plot_return_period( spec_c=self.dico_subplot_evts[ind_evt]+np.array([0,1]), ind_evt=ind_evt )
            else:
                subax_attribgmt, bbox_2 = self.plot_gmt( spec_c=self.dico_subplot_evts[ind_evt]+np.array([0,1]), ind_evt=ind_evt, ranges=[50, 90, 95], alpha_ranges=[0.5,0.3,0.15] )
            
            # plot map of region of event
            subax_reg, pmesh_reg = self.plt_reg( ax_map=ax_map, spec_c=self.dico_subplot_evts[ind_evt], obs_selecttime=obs_selecttime)
            pos = subax_reg.get_position()
            subax_reg.set_position([ pos.x0-(self.factor_sizes_reg-1)*pos.width-self.shift_subplot_reg, \
                                    pos.y0-0.5*(self.factor_sizes_reg-1)*pos.height + self.dico_shifts_subplots[self.dico_subplot_evts[ind_evt][0]], \
                                    pos.width*self.factor_sizes_reg, \
                                    pos.height*self.factor_sizes_reg])
            # adding colorbar, with size matching axis
            pos_new = subax_reg.get_position()
            cax = self.fig.add_axes([pos_new.x0-0.016-self.shift_subplot_reg, pos_new.y0 + 0.1*pos_new.height, 0.0075, 0.8*pos_new.height])
            cbar = plt.colorbar(pmesh_reg, cax=cax, location='left', orientation ='vertical', extend='both', extendfrac=0.025)
            cbar.ax.tick_params(labelsize=self.fontsizes['reg_cbar_ticks'])
            cbar.ax.set_xlabel('T ('+u'\u00B0C'+')', size=self.fontsizes['reg_cbar_label'])
            t = subax_reg.text(x=-0.20, y=1.0, s='('+self.letter_panel+')', transform=subax_reg.transAxes, fontdict={'size':self.fontsizes['letter_panel']}, weight='bold',\
                               verticalalignment='center', horizontalalignment='center', rotation_mode='anchor',\
                               bbox={'boxstyle':'round', 'facecolor':self.dico_cols['letter_panel_face'], 'edgecolor':self.dico_cols['letter_panel_edge'], 'alpha':1})

            # creating rectangle for each case
            bbox_1 = subax_reg.get_position()
            # slightly increase the very tight bounds:
            xpad_left, xpad_right = 0.225 * bbox_2.width, 0.175 * bbox_2.width
            ypad_low, ypad_up = 0.215 * bbox_2.height, 0.025 * bbox_2.height # 0.16, 0.125
            rect = plt.Rectangle((bbox_1.x0-xpad_left, bbox_2.y0+self.dico_shifts_subplots[self.dico_subplot_evts[ind_evt][0]]-ypad_low),\
                                 bbox_2.x1-bbox_1.x0 + xpad_left + xpad_right,\
                                 bbox_2.height + ypad_low + ypad_up,\
                                 edgecolor=(0.5,0.5,0.5), facecolor=(1.0,1.0,1.0,0.0), linewidth=self.dico_lw['rect_cases'], zorder=-1000) #edgecolor=(0.5,0.5,0.5), facecolor=(0.9,0.9,0.9,0.5)
            self.fig.add_artist(rect)

            # plot arrow from initial region to map of region of event
            posA = np.array( [self.central_longitude_evt, 0.5*sum(self.lat_borders_evt)] )
            posB = np.array( [-180 + (np.sum(self.width_ratios[:self.dico_subplot_evts[ind_evt][1]]) + 0.75*self.width_ratios[self.dico_subplot_evts[ind_evt][1]]) * 360/np.sum(self.width_ratios),\
                              90 - (self.dico_subplot_evts[ind_evt][0]==2)*180] )
            if posB[0] < 0: # due to the projection, and arrow up to the edge
                posB[0] = -179
            if arrow_or_label in ['arrow','both']:
                ax_map.add_patch( FancyArrowPatch(posA=posA+np.sign(posB-posA)*np.array([5,2.5]), posB=posB, mutation_scale=25, arrowstyle=self.arrow_style,\
                                                  edgecolor=(0.50,0.50,0.50,0.50), facecolor=(0.50,0.50,0.50,0.50), lw=self.dico_lw['arrow'], transform=ccrs.PlateCarree()) )#edgecolor=(0.25,0.25,0.25,1.0), facecolor=(0.3,0.3,0.3,0.85)
            if arrow_or_label in ['label','both']:
                t = ax_map.text(x=posA[0], y=posA[1], s='('+self.letter_panel+')', transform=ccrs.PlateCarree(), fontdict={'size':self.fontsizes['letter_panel']}, weight='bold',\
                                va='center', ha='center', rotation_mode='anchor',\
                                bbox={'boxstyle':'round', 'facecolor':self.dico_cols['letter_panel_face'], 'edgecolor':self.dico_cols['letter_panel_edge'], 'alpha':1})

        # save
        self.fig.savefig( os.path.join( path_save, 'figure1_v4.png' ), dpi=300 )
        self.fig.savefig( os.path.join( path_save, 'figure1_v4.pdf' ) )
        return self.fig
    #--------------------------------------------------------------


    #--------------------------------------------------------------
    # GLOBAL MAP
    def plot_map_country_events( self, ax):#, cbaxes ):
        # defining countries
        countries = regionmask.defined_regions.natural_earth_v5_0_0.countries_110
        
        # preparing grid
        lon = np.arange(-180,180, 0.25)
        lat = np.arange(-90, 90, 0.25)
        mask = countries.mask(lon, lat)
        
        # counting events
        mask_new = self.counting_events_countries(countries, mask)

        # preparing colors for maps
        norm_map = plcol.Normalize(vmin=mask_new.min(), vmax=mask_new.max() )
        if False:
            levels_map = [-0.5, 0.5, 1.5, 5.5, 15.5]
            levels_cbar = [0, 1, 3.5, 10.5]
            text_cbar = ['0', '1', '2-5', '6+']
            c = 0.995
            ticks_cbar = [(0,(1.0*c,1.0*c,1.0*c)), (0.5/15,(0.90*c,0.90*c,0.90*c)), (1/15,(0.8*c,0.8*c,0.8*c)), (3.5/15,(0.5*c,0.5*c,0.5*c)), (15/15,(0.0*c,0.0*c,0.0*c))]
        else:
            levels_map = [-0.5, 0.5, 2.5, 4.5, 6.5, 8.5, 10.5, 15.5]
            levels_cbar = [0, 1.5, 3.5, 5.5, 7.5, 9.5, 13]
            text_cbar = ['0', '1-2', '3-4', '5-6', '7-8', '9-10', '11+']
            c0, c1, c2 = np.array([248/255, 248/255, 254/255]), np.array([125/255, 100/255, 227/255]), np.array([118/255, 27/255, 122/255])
            ticks_cbar = [(0,c0), (1.5/15,c0+1/3*(c1-c0)), (3.5/15, c0+2/3*(c1-c0)),\
                          (5.5/15,c1, c1+1/3*(c2-c1)), (9.5/15,c1+2/3*(c2-c1)),\
                          (15/15,c2)]
            #c = 0.995
            #ticks_cbar = [(0,(1.0*c,1.0*c,1.0*c)), (1.5/15,(0.85*c,0.85*c,0.85*c)), (3.5/15,(0.7*c,0.7*c,0.7*c)),\
            #              (5.5/15,(0.55*c,0.55*c,0.55*c)), (7.5/15,(0.4*c,0.4*c,0.4*c)), (9.5/15,(0.2*c,0.2*c,0.2*c)),\
            #              (15/15,(0.0*c,0.0*c,0.0*c))]
        cmap_map = plcol.LinearSegmentedColormap.from_list('custom', ticks_cbar)

        # ploting
        countries.plot(ax=ax, add_label=False)
        pmesh = ax.contourf(lon, lat, mask_new, transform=ccrs.PlateCarree(), cmap=cmap_map, levels=levels_map)
        _ = plt.setp(ax.spines.values(), linewidth=self.dico_lw['global_map'])

        # polishing
        cbar = plt.colorbar(pmesh, orientation ='vertical', fraction=0.023, pad=0.005, shrink=0.75 )
        period_emdat = str(int(self.emdat['Start Year'].min().values)) + '-' + str(int(self.emdat['Start Year'].max().values))
        cbar.set_label( 'Number of reported events over '+period_emdat, size=self.fontsizes['global_cbar_label'])
        cbar.ax.tick_params(labelsize=self.fontsizes['global_cbar_ticks'])
        cbar.set_ticks( levels_cbar )
        cbar.set_ticklabels( text_cbar )
        
        
    def counting_events_countries( self, countries, mask ):
        # Malta to Cyprus: most adequate country in regionmask, although not ideal from a geopolitical perspective
        dico_countries = {cou:0 for cou in countries.names}
        for ind_evt in self.emdat.index.values:
            # identifying country
            cou = str(self.emdat['Country'].sel(index=ind_evt).values)
            if cou not in dico_countries:
                cou = dico_add_emdatcountries_regionmask[cou]
            else:
                cou = [cou]

            # increasing value by 1
            for c in cou:
                dico_countries[c] += 1

        # affecting in mask
        mask_new = np.nan * xr.ones_like( mask ) 
        for cou in dico_countries:
            mask_new = xr.where( mask == countries.region_ids[cou], dico_countries[cou], mask_new)
        return mask_new
    #--------------------------------------------------------------

        
    #--------------------------------------------------------------
    # PLOT MAP OF EVENT (ADAPTED FROM CLASS ON EVENT DEFINITION)
    def id_map_event(self, obs_selecttime, type_reg, margin_lat=1.5, margin_lon=1.5, set_ratio_lon_lat=None):
        # preparing mask: whole country
        if type_reg == 'country':
            list_all_subisos = np.hstack( [self.evt_id.dict_geobounds[iso] for iso in self.evt_id.event_iso] )
            spareg = gpd.GeoSeries( unary_union(self.evt_id.geobounds.loc[list_all_subisos]['geometry']) )
            mask = self.evt_id.mask_reg( spareg, obs_selecttime[self.window].lon, obs_selecttime[self.window].lat )
        elif type_reg == 'event':
            spareg = self.evt_id.event_region
            mask = self.evt_id.mask_reg( spareg, obs_selecttime[self.window].lon, obs_selecttime[self.window].lat )
        mask = mask.where( mask == 0, 1 ) # adapting to have all grid cells for correct min/max
        tmp = mask.where( mask>0, drop=True )
        lon_plot, lat_plot = tmp.lon.values, tmp.lat.values
        lon_plot[np.where(lon_plot>180)] -= 360

        # preparing central longitude and borders of map: longitude done in dirty way, but couldnt find better approach
        # SUPER MESSY PART THERE
        lat_borders = np.array( [np.min(lat_plot)-margin_lat, np.max(lat_plot)+margin_lat] )
        test_lon180 = 180 - np.max( [np.abs(np.min(lon_plot)),np.max(lon_plot)] ) < np.max( [np.abs(np.min(lon_plot)),np.max(lon_plot)] )
        if test_lon180:
            tmp_lon = np.copy(lon_plot)
            tmp_lon[np.where(lon_plot<0)] += 360
            lon_borders = np.array( [np.min(tmp_lon)-margin_lon, np.max(tmp_lon)+margin_lon] )
            central_longitude = np.mean(lon_borders)
            lon_borders[np.where(lon_borders>180)] -= 360
        else:
            lon_borders = np.array( [np.min(lon_plot+360)-margin_lon, np.max(lon_plot+360)+margin_lon] )
            central_longitude = np.mean(lon_borders)
        if lon_borders[0] > lon_borders[1]:
            lon_borders[0] -= 360
        if lon_borders[0] > 180:
            lon_borders[0] -= 360
        if lon_borders[1] > 180:
            lon_borders[0] -= 360
        #if (lon_borders[0] > 360) and (lon_borders[1]> 360):
        #    lon_borders[0] -= 360
        #    lon_borders[1] -= 360
        # experimental feature for figure 1 of paper
        if set_ratio_lon_lat is not None: 
            lo = lon_borders[1] - lon_borders[0]
            la = lat_borders[1] - lat_borders[0]
            if set_ratio_lon_lat < 1:
                lat_borders[0] -= 0.5 * ( lo/set_ratio_lon_lat - la )
                lat_borders[1] += 0.5 * ( lo/set_ratio_lon_lat - la )
            else:
                lon_borders[0] -= 0.5 * ( set_ratio_lon_lat*la - lo )
                lon_borders[1] += 0.5 * ( set_ratio_lon_lat*la - lo )
        if central_longitude > 180:
            central_longitude -= 360
        if (central_longitude < 0) and (obs_selecttime[self.window].lon.min() > 0):
            central_longitude += 360 # case USA on CMIP6-ng grid
        return spareg, mask, central_longitude, lon_borders, lat_borders

    
    def plt_reg(self, ax_map, spec_c, obs_selecttime):
        # centering projection (in case of countries stretching over -180 East: USA, New Zealand, etc)
        ax_window = plt.subplot( self.spec[spec_c[0],spec_c[1]], projection=ccrs.Robinson(central_longitude=self.central_longitude) )
        ax_window.coastlines()

        # ploting whole country
        reg = regionmask.Regions([self.country.item()])
        reg.plot_regions(ax=ax_window, add_label=False, line_kws={'color':self.dico_cols['reg_map_country'],'linestyle':'-','lw':self.dico_lw['reg_map_country']} )
        
        # ploting selected region
        reg = regionmask.Regions([self.evt_id.event_region.item()])
        reg.plot_regions(ax=ax_window, add_label=False, line_kws={'color':self.dico_cols['reg_map_event'], 'linestyle':'-', 'lw':self.dico_lw['reg_map_reg']})

        # identifying correct variable
        dt = obs_selecttime[self.window][self.evt_id.var_in_data(obs_selecttime[self.window])].sel( time=self.evt_id.event_period['Start']['Year'] ).drop('time')
        if 'member' in dt.coords:
            dt = dt.mean('member')
        data_map = dt.compute()

        # ploting
        pmesh = ax_window.pcolormesh(obs_selecttime[self.window]['lon'], \
                                     obs_selecttime[self.window]['lat'], \
                                     xr.where(self.mask>0, data_map, np.nan), cmap='Reds', \
                                     transform=ccrs.PlateCarree(), rasterized=True , vmin=xr.where( self.mask>0, data_map, np.nan).min(), vmax=xr.where( self.mask>0, data_map, np.nan).max() )
        ax_window.set_extent( (self.lon_borders[0], self.lon_borders[1], self.lat_borders[0], self.lat_borders[1]), crs=ccrs.PlateCarree())

        # text for period of event
        tmp = {}
        for tm in ['Start', 'End']:
            if 'time_'+tm in self.evt_id.warnings:
                tmp[tm] = str(self.evt_id.event_period[tm]['Year']) + '.' + str(self.evt_id.event_period[tm]['Month'])
            else:
                tmp[tm] = str(self.evt_id.event_period[tm]['Year']) + '.' + str(self.evt_id.event_period[tm]['Month']) + '.' + str(self.evt_id.event_period[tm]['Day'])
        textstr = "Reported period: " + "\n" + tmp['Start'] + ' - ' + tmp['End']
        t2 = ax_window.text(x=0.5, y=-0.125, s=textstr, transform=ax_window.transAxes, fontdict={'size':self.fontsizes['attrib_txt_period']},\
                            verticalalignment='center', horizontalalignment='center', rotation_mode='anchor', bbox={'boxstyle':'round', 'facecolor':(0.95,0.95,0.95,1), 'alpha':1})
        
        # adding title
        #ttl0 = '$T_{'+self.window+'}$ of '+self.evt_id.name_data
        #ax_window.text(-0.175, 0.55, s=ttl0, fontdict={'size':0.9*self.fontsize}, va='bottom', ha='center', rotation='vertical', rotation_mode='anchor', transform=ax_window.transAxes)

        # adding colorbar after changes in position of subplot
        return ax_window, pmesh
    #--------------------------------------------------------------


    #--------------------------------------------------------------
    # PLOT OF RETURN PERIOD
    @staticmethod
    def eval_I_at_return_period(return_periods, distrib, params, yr, label, dico_q_confid):
        p = xr.DataArray( 1 - 1 / return_periods, coords={'return_period':return_periods}, dims=('return_period',)).expand_dims('bootstrap',axis=1)
        # calculating probabilities to have this value or below
        if distrib == 'GEV':
            tmp1 = ss.genextreme.ppf(p,\
                                     loc=params['loc'].sel(time=yr, label=label).expand_dims('return_period',axis=0),\
                                     scale=params['scale'].sel(time=yr, label=label).expand_dims('return_period',axis=0),\
                                     c=-params['shape'].sel(time=yr, label=label).expand_dims('return_period',axis=0) )
        elif distrib == 'gaussian':
            tmp1 = ss.norm.ppf(p,\
                               loc=params['loc'].sel(time=yr, label=label),\
                               scale=params['scale'].sel(time=yr, label=label) )
        elif distrib == 'poisson':
            tmp1 = ss.poisson.ppf(p,\
                                  loc=params['loc'].sel(time=yr, label=label),\
                                  mu=params['mu'].sel(time=yr, label=label) )
        elif distrib == 'GPD':
            tmp1 = ss.genpareto.ppf(p,\
                                     loc=params['loc'].sel(time=yr, label=label),\
                                     scale=params['scale'].sel(time=yr, label=label),\
                                     c=params['shape'].sel(time=yr, label=label) )
        elif distrib == 'skewnorm':
            tmp1 = ss.skewnorm.ppf(p,\
                                   loc=self.parameters[window]['loc'].sel(time=yr, label=label),\
                                   scale=self.parameters[window]['scale'].sel(time=yr, label=label),\
                                   a=self.parameters[window]['shape'].sel(time=yr, label=label) )
    
        # finalize dataset
        int_at_RP = xr.Dataset()
        int_at_RP['values'] = xr.DataArray( tmp1, coords={'return_period':return_periods, 'bootstrap':params.bootstrap}, dims=('return_period', 'bootstrap',) )
        
        # calculating mean, median, confidence interval
        int_at_RP['mean'] = int_at_RP['values'].mean('bootstrap')
        int_at_RP['median'] = int_at_RP['values'].median('bootstrap')
        for q in dico_q_confid:
            int_at_RP[q] = int_at_RP['values'].quantile(q=dico_q_confid[q], dim='bootstrap').drop('quantile')
        return int_at_RP

    @staticmethod
    def sci_notation(number, sig_fig=2, opt_integer=True):
        if np.isnan(number) or np.isinf(number):
            return "$+\infty$"
        else:
            ret_string = "{0:.{1:d}e}".format(number, sig_fig)
            a, b = ret_string.split("e")
            # remove leading "+" and strip leading zeros
            b = int(b)
            if b <= sig_fig:
                if opt_integer and (int(float(a)) == float(a)):
                    return "$" + str(int(float(a)*10**b)) + "$"
                else:
                    return "$" + str(float(a)*10**b) + "$"
            else:
                if opt_integer and (int(float(a)) == float(a)):
                    return "$" + str(int(float(a))) + ".10^{" + str(b) + "}$"
                else:
                    return "$" + a + ".10^{" + str(b) + "}$"

    @staticmethod
    def func_text_arrow( val ):
        med = sci_notation(val['median'], sig_fig=2)
        bot = sci_notation(val['confid_bottom'], sig_fig=2)
        upp = sci_notation(val['confid_upper'], sig_fig=2)
        if upp == '$+\\infty$':
            return med +' [' + bot + '; ' + upp + '['
        else:
            return med +' [' + bot + '; ' + upp + ']'

    def plot_return_period( self, spec_c, ind_evt ):
        bounds_power = [0, 4]
        
        # creating subplot
        ax_window = plt.subplot( self.spec[spec_c[0],spec_c[1]] )

        # preparation
        xticks = [10**i for i in np.arange(bounds_power[0],bounds_power[1]+1)]
        return_periods = np.logspace( bounds_power[0], bounds_power[1], int(1e5) )
        distrib = self.evt_fits[self.name_data].best_fit[self.window]['distrib']
        params = self.evt_fits[self.name_data].parameters[self.window]
        yr = self.evt_fits[self.name_data].event_year
        dico_q_confid = self.evt_fits[self.name_data].dico_q_confid
        int_RP_withCC = self.eval_I_at_return_period(return_periods, distrib, params, yr, 'with_CC', dico_q_confid)
        int_RP_withoutCC = self.eval_I_at_return_period(return_periods, distrib, params, yr, 'without_CC', dico_q_confid)
        
        # plot
        ax_window.plot( return_periods, int_RP_withoutCC['median'], color=self.dico_cols['without_CC'], lw=self.dico_lw['return_wwo'], label='Without: ' )
        ax_window.fill_between( return_periods, int_RP_withoutCC['confid_bottom'], int_RP_withoutCC['confid_upper'], color=self.dico_cols['without_CC'], alpha=0.5 )
        ax_window.plot( return_periods, int_RP_withCC['median'], color=self.dico_cols['with_CC'], lw=self.dico_lw['return_wwo'], label='With: ' )
        ax_window.fill_between( return_periods, int_RP_withCC['confid_bottom'], int_RP_withCC['confid_upper'], color=self.dico_cols['with_CC'], alpha=0.5 )
        ax_window.set_xscale('log')
        ax_window.set_xlim(return_periods[0], return_periods[-1])
        ax_window.set_ylim(int_RP_withoutCC['confid_bottom'].min(), int_RP_withCC['confid_upper'].max())
        
        # preparing lines & arrows
        #pr_text = func_text_arrow(self.evt_fits[self.name_data].PR[self.window])
        #i_text = func_text_arrow(self.evt_fits[self.name_data].I[self.window])
        p_withCC = self.evt_fits[self.name_data].probabilities[self.window]['median'].sel(label='with_CC')
        rp_withCC = 1 / p_withCC
        if np.isinf(rp_withCC) or np.isnan(rp_withCC) or (rp_withCC > ax_window.get_xlim()[1]):
            rp_withCC = ax_window.get_xlim()[1]
        p_withoutCC = self.evt_fits[self.name_data].probabilities[self.window]['median'].sel(label='without_CC')
        rp_withoutCC = 1 / p_withoutCC
        if np.isinf(rp_withoutCC) or np.isnan(rp_withoutCC) or (rp_withoutCC > ax_window.get_xlim()[1]):
            rp_withoutCC = ax_window.get_xlim()[1]
        tmp_i_withCC = self.evt_fits[self.name_data].intensities[self.window]['median'].sel(label='with_CC')
        tmp_i_withoutCC = self.evt_fits[self.name_data].intensities[self.window]['median'].sel(label='without_CC')
        evt_level = self.evt_fits[self.name_data].event_level[self.window]
        i_withCC = evt_level
        i_withoutCC = evt_level + tmp_i_withoutCC - tmp_i_withCC
        
        # plotting line
        _ = ax_window.vlines( x=[rp_withCC], ymin=ax_window.get_ylim()[0], ymax=i_withCC, color=self.dico_cols['event'], lw=self.dico_lw['return_lines'], ls='--' )
        _ = ax_window.hlines( y=[i_withCC], xmin=ax_window.get_xlim()[0], xmax=rp_withCC, color=self.dico_cols['event'], lw=self.dico_lw['return_lines'], ls='--' )
        
        # plotting arrows
        _ = ax_window.annotate(text='', xytext=(rp_withoutCC, i_withCC), xy=(rp_withCC, i_withCC), arrowprops=dict(facecolor='red', edgecolor=(0,0,0,0), shrink=0., lw=self.dico_lw['return_arrows']))
        _ = ax_window.annotate(text='', xytext=(rp_withCC, i_withoutCC), xy=(rp_withCC, i_withCC), arrowprops=dict(facecolor='red', edgecolor=(0,0,0,0), shrink=0., lw=self.dico_lw['return_arrows']))
        
        # legend & ticks
        handles, labels = ax_window.get_legend_handles_labels()
        legend = ax_window.legend([handles[1],handles[0]], [labels[1],labels[0]], loc='lower right', frameon=False, markerfirst=False, borderaxespad=0,\
                                  title="Climate change:", prop={'size':self.fontsizes['attrib_legend']}, title_fontsize=self.fontsizes['attrib_legend_title'])
        # the normal way to change the fontsize of ticks is as usually bad. works depending on which ind_evt... doing that the hardcore way.
        for tick in ax_window.xaxis.get_major_ticks():
            tick.label.set_fontsize(self.fontsizes['attrib_axis_ticks']) 
        for tick in ax_window.yaxis.get_major_ticks():
            tick.label.set_fontsize(self.fontsizes['attrib_axis_ticks']) 
        #_ = plt.xticks( fontsize=self.fontsizes['attrib_axis_ticks'] ) # size
        #_ = plt.yticks( fontsize=self.fontsizes['attrib_axis_ticks'] ) # size
        ax_window.set_xlabel( 'Return period', size=self.fontsizes['attrib_axis_labels'], labelpad=-5 )
        ax_window.set_ylabel( 'Intensity ('+u'\u00B0C'+')', size=self.fontsizes['attrib_axis_labels'], labelpad=-1 )
        plt.draw()
        pos = ax_window.get_position()
        ax_window.set_position([pos.x0, pos.y0 + self.dico_shifts_subplots[self.dico_subplot_evts[ind_evt][0]], pos.width, pos.height])
        pos = ax_window.get_position()
        return ax_window, pos
    #--------------------------------------------------------------



    #--------------------------------------------------------------
    # PLOT GMT
    def plot_gmt( self, spec_c, ind_evt, ranges, alpha_ranges, bar_or_arrow='arrow' ):
        # preparing gmt
        gmt_full = self.evt_fits[self.name_data].data_gmt
        rescale_pi = (gmt_full.sel(time=slice(1961,1990)).mean() - 0.36)# IPCC AR6 Chapter2, Cross Chapter Box 2.3, Table 1
        gmt_range_full = xr.DataArray(np.linspace(rescale_pi - 0.05*(gmt_full.max()-gmt_full.min()), gmt_full.max() + 0.05*(gmt_full.max()-gmt_full.min()), gmt_full.time.size), coords={'time':gmt_full.time})
        
        # for plots
        gmt = gmt_full - rescale_pi
        gmt_range = gmt_range_full - rescale_pi
        obs = self.evt_fits[self.name_data].data_obs[self.window]
        yr = self.evt_fits[self.name_data].event_year
        
        # preparing subplot
        ax = plt.subplot( self.spec[spec_c[0],spec_c[1]] )

        # scatter plot
        plt_obs = ax.scatter( gmt, obs, s=10, color=self.dico_cols['event'], zorder=1000, label=self.name_data )
        ax.scatter( gmt.sel(time=yr), obs.sel(time=yr), s=60, color=self.dico_cols['event'], zorder=1000 )

        # ploting each level of range
        ranges.sort(reverse=True)
        alpha_ranges.sort(reverse=False)
        plt_distrib = []
        for ilvl, level_range in enumerate(ranges):
            self.evt_fits[self.name_data].dico_q_range ={'low':(50-level_range/2)/100, 'high':(50+level_range/2)/100}
            
            # preparing distribution
            self.evt_fits[self.name_data].calc_params_bootstrap( predictor=gmt_range_full, label='plot_distrib_'+str(level_range) )
            if self.evt_fits[self.name_data].best_fit[self.window]['distrib'] != 'GEV':
                raise Exception("The rest has been simplified with only GEV in mind, for the plots. Anyway, the next version of this work will be much easier with the distributions, cf current devs on MESMER-X")
            
            # for each quantile, values for all bootstrapped member
            values = {}
            for q in self.evt_fits[self.name_data].dico_q_range:
                params = {'loc':self.evt_fits[self.name_data].parameters[self.window]['loc'].sel(label='plot_distrib_'+str(level_range)),\
                          'scale':self.evt_fits[self.name_data].parameters[self.window]['scale'].sel(label='plot_distrib_'+str(level_range)),\
                          'c':-self.evt_fits[self.name_data].parameters[self.window]['shape'].sel(label='plot_distrib_'+str(level_range))}
                # if run multiple times, may add several times same label (TO DO: in fcts_support_training, edit to rewrite data if same label)
                for pp in params:
                    if params[pp].label.size > 1:
                        params[pp] = params[pp].isel(label=-1)
                values[q] = ss.genextreme.ppf(q=1 - self.evt_fits[self.name_data].dico_q_range[q], loc=params['loc'], scale=params['scale'], c=params['c'])
            values['median'] = ss.genextreme.ppf(q=0.5, loc=params['loc'], scale=params['scale'], c=params['c'])
            # calculate uncertainty range for all
            intensities = {}
            for q in values:
                for u in self.evt_fits[self.name_data].dico_q_confid:
                    intensities[q+'_u'+u] = np.nanpercentile( values[q], q=100*self.evt_fits[self.name_data].dico_q_confid[u], axis=0)
                intensities[q+'_umedian'] = np.median(values[q], axis=0)
    
            # fill between for ranges
            plt_distrib.append( ax.fill_between(gmt_range, intensities['low_umedian'], intensities['high_umedian'],\
                                                facecolor=self.dico_cols['attrib'], edgecolor=None, lw=0, alpha=alpha_ranges[ilvl], label=str(level_range)+'%') )# alpha=np.sqrt(1-level_range/100)

        # plotting median
        tmp = ax.plot( gmt_range, intensities['median_umedian'], lw=self.dico_lw['gmt_central'], color=self.dico_cols['attrib_median'], label='median')
        plt_distrib.append( tmp[0] )
        
        # improving plot
        xl = gmt_range.min(), gmt_range.max()
        ax.set_xlim(xl)
        yl = ax.get_ylim()[0] - 0.15*(ax.get_ylim()[1]-ax.get_ylim()[0]), ax.get_ylim()[1] + 0.15*(ax.get_ylim()[1]-ax.get_ylim()[0])
        ax.set_ylim(yl)
        if gmt.sel(time=yr) > xl[0] + 0.80*(xl[1]-xl[0]):
            t = ax.text(x=gmt.sel(time=yr).values-0.025*(xl[1]-xl[0]), y=obs.sel(time=yr).values+0.025*(yl[1]-yl[0]), s='Event',\
                        fontdict={'size':self.fontsizes['gmt_text']}, verticalalignment='bottom', horizontalalignment='right', rotation_mode='anchor')
        else:
            t = ax.text(x=gmt.sel(time=yr).values+0.025*(xl[1]-xl[0]), y=obs.sel(time=yr).values+0.025*(yl[1]-yl[0]), s='Event',\
                        fontdict={'size':self.fontsizes['gmt_text']}, verticalalignment='bottom', horizontalalignment='left', rotation_mode='anchor')
        
        # event
        ax.vlines( x=gmt.sel(time=yr), ymin=yl[0], ymax=obs.sel(time=yr), ls='--', color=self.dico_cols['event'], lw=self.dico_lw['gmt_event'] )
        ax.hlines( y=obs.sel(time=yr), xmin=xl[0], xmax=gmt.sel(time=yr), ls='--', color=self.dico_cols['event'], lw=self.dico_lw['gmt_event'] )
        
        # legends
        plot_1st_legend = plt_distrib[::-1]
        lgd1 = plt.legend(plot_1st_legend, [pp.get_label() for pp in plot_1st_legend], loc='lower right',\
                          prop={'size':self.fontsizes['gmt_legend']}, frameon=True, markerfirst=False, ncols=2, handletextpad=0.1, title="Distribution:", title_fontsize=self.fontsizes['gmt_legend_title']) # , borderaxespad=0
        ax.add_artist(lgd1)
        plot_2nd_legend = [plt_obs]
        lgd2 = plt.legend(plot_2nd_legend, [pp.get_label() for pp in plot_2nd_legend], loc='upper left',\
                          prop={'size':self.fontsizes['gmt_legend']}, frameon=True, markerfirst=False, handletextpad=0.1)
        ax.add_artist(lgd2)
        
        # finish plot
        _ = plt.xticks( size=self.fontsizes['gmt_axis_ticks'] )
        _ = plt.yticks( size=self.fontsizes['gmt_axis_ticks'] )
        ax.set_xlabel( 'Change in Global Mean Temperature\nwith reference to 1850-1900 ('+u'\u00B0C'+')', size=self.fontsizes['gmt_axis_labels'] )
        ax.set_ylabel( 'Intensity T ('+u'\u00B0C'+')', size=self.fontsizes['gmt_axis_labels'], labelpad=-1 )
        pos = ax.get_position()
        ax.set_position([pos.x0, pos.y0 + self.dico_shifts_subplots[self.dico_subplot_evts[ind_evt][0]],\
                         pos.width - (self.shift_axbars_gmt_int + self.shift_axbars_gmt_prb + 2*self.width_axbars_gmt), pos.height])
        
        # adding bar for intensities
        pos_new = ax.get_position()
        ax_intensity = self.fig.add_axes([pos_new.x1+self.shift_axbars_gmt_int, pos_new.y0, self.width_axbars_gmt, pos_new.height])
        ax_intensity.set_xlim(0,1)
        ax_intensity.set_ylim(yl)
        its = self.evt_fits[self.name_data].intensities[self.window]
        # plot bars
        if bar_or_arrow == 'arrow':
            pass
        else:
            _ = ax_intensity.bar( x=0.5, bottom=yl[0], height=its['median'].sel(label='without_CC')-yl[0], width=1, color=self.dico_cols['without_CC'] )
        if False:
            _ = ax_intensity.hlines(y=its['confid_bottom'].sel(label='without_CC'), xmin=0, xmax=1,\
                                    ls='--', lw=self.dico_lw['gmt_uncertainties'], color=self.dico_cols['uncertainties'] )
            _ = ax_intensity.hlines(y=its['confid_upper'].sel(label='without_CC'), xmin=0, xmax=1,\
                                    ls='--', lw=self.dico_lw['gmt_uncertainties'], color=self.dico_cols['uncertainties'] )
        _ = ax_intensity.hlines(y=its['median'].sel(label='without_CC'), xmin=0, xmax=1,\
                                ls='-', lw=self.dico_lw['gmt_median_ip'], color=self.dico_cols['counter_event'], zorder=0 )
        if bar_or_arrow == 'arrow':
            ax_intensity.add_patch( FancyArrowPatch(posA=(0.5,its['median'].sel(label='without_CC').values), posB=(0.5,its['median'].sel(label='with_CC').values), mutation_scale=25, arrowstyle=self.arrow_style,\
                                              edgecolor=self.dico_cols['attrib'], facecolor=self.dico_cols['attrib'], lw=self.dico_lw['arrow']) )
        else:
            _ = ax_intensity.bar(x=0.5, bottom=its['median'].sel(label='without_CC'),\
                                 height=its['median'].sel(label='with_CC')-its['median'].sel(label='without_CC'), width=1, color=self.dico_cols['with_CC'] )
        if False:
            _ = ax_intensity.hlines(y=its['confid_bottom'].sel(label='with_CC'), xmin=0, xmax=1,\
                                    ls='--', lw=self.dico_lw['gmt_uncertainties'], color=self.dico_cols['uncertainties'] )
            _ = ax_intensity.hlines(y=its['confid_upper'].sel(label='with_CC'), xmin=0, xmax=1,\
                                    ls='--', lw=self.dico_lw['gmt_uncertainties'], color=self.dico_cols['uncertainties'] )
        _ = ax_intensity.hlines(y=its['median'].sel(label='with_CC'), xmin=0, xmax=1,\
                                ls='-', lw=self.dico_lw['gmt_median_ip'], color=self.dico_cols['event'], zorder=0 )
        ax_intensity.axes.get_xaxis().set_visible(False)
        ax_intensity.axes.get_yaxis().set_visible(False)
        _ = ax_intensity.set_title('T ('+u'\u00B0C'+')', size=self.fontsizes['gmt_title'], y=-0.20, rotation=90)#, y=-0.15
        
        # adding bar for probabilities
        ax_probabilities = self.fig.add_axes([pos_new.x1+self.shift_axbars_gmt_int + self.shift_axbars_gmt_prb + self.width_axbars_gmt, pos_new.y0,\
                                              self.width_axbars_gmt, pos_new.height])
        ax_probabilities.set_xlim(0,1)
        prb = self.evt_fits[self.name_data].probabilities[self.window]
        # plot bars
        if bar_or_arrow == 'arrow':
            pass
        else:
            _ = ax_probabilities.bar( x=0.5, bottom=0, height=prb['median'].sel(label='without_CC'), width=1, color=self.dico_cols['without_CC'] )
        if False:
            _ = ax_probabilities.hlines(y=prb['confid_bottom'].sel(label='without_CC'), xmin=0, xmax=1,\
                                        ls='--', lw=self.dico_lw['gmt_uncertainties'], color=self.dico_cols['uncertainties'] )
            _ = ax_probabilities.hlines(y=prb['confid_upper'].sel(label='without_CC'), xmin=0, xmax=1,\
                                        ls='--', lw=self.dico_lw['gmt_uncertainties'], color=self.dico_cols['uncertainties'] )
        _ = ax_probabilities.hlines(y=prb['median'].sel(label='without_CC'), xmin=0, xmax=1,\
                                ls='-', lw=self.dico_lw['gmt_median_ip'], color=self.dico_cols['counter_event'], zorder=0 )
        if bar_or_arrow == 'arrow':
            ax_probabilities.add_patch( FancyArrowPatch(posA=(0.5,prb['median'].sel(label='without_CC').values), posB=(0.5,prb['median'].sel(label='with_CC').values), mutation_scale=25, arrowstyle=self.arrow_style,\
                                                        edgecolor=self.dico_cols['attrib'], facecolor=self.dico_cols['attrib'], lw=self.dico_lw['arrow'] ) )
        else:
            _ = ax_probabilities.bar(x=0.5, bottom=prb['median'].sel(label='without_CC'),\
                                     height=prb['median'].sel(label='with_CC')-prb['median'].sel(label='without_CC'), width=1, color=self.dico_cols['with_CC'], zorder=0 )
        if False:
            _ = ax_probabilities.hlines(y=prb['confid_bottom'].sel(label='with_CC'), xmin=0, xmax=1,\
                                        ls='--', lw=self.dico_lw['gmt_uncertainties'], color=self.dico_cols['uncertainties'] )
            _ = ax_probabilities.hlines(y=prb['confid_upper'].sel(label='with_CC'), xmin=0, xmax=1,\
                                        ls='--', lw=self.dico_lw['gmt_uncertainties'], color=self.dico_cols['uncertainties'] )
        _ = ax_probabilities.hlines(y=prb['median'].sel(label='with_CC'), xmin=0, xmax=1,\
                                ls='-', lw=self.dico_lw['gmt_median_ip'], color=self.dico_cols['event'] )

        # finishing the plot
        yl = [np.max([1.e-4,prb['confid_bottom'].sel(label='without_CC')]), prb['confid_upper'].sel(label='with_CC')]
        yl[1] += 0.025*(yl[1]-yl[0])
        _ = ax_probabilities.set_ylim(yl)
        ax_probabilities.set_yscale('log')
        # labels=['$10^'+str(-int(np.log10(s)))+'$' for s in ax_probabilities.get_yticks()]
        labels=[int(10**(-int(np.log10(s)))) for s in ax_probabilities.get_yticks()]
        ax_probabilities.set_yticks(ticks=ax_probabilities.get_yticks(), labels=labels, size=self.fontsizes['gmt_axis_ticks_bars'] )
        ax_probabilities.axes.get_xaxis().set_visible(False)
        ax_probabilities.yaxis.tick_right()
        _ = ax_probabilities.set_ylim(yl)
        _ = ax_probabilities.set_title('Return\nperiod\n(year)', size=self.fontsizes['gmt_title'], y=-0.20, rotation=90)
        return ax, pos
    #--------------------------------------------------------------
#---------------------------------------------------------------------------------------------------------------------------
#---------------------------------------------------------------------------------------------------------------------------




















#---------------------------------------------------------------------------------------------------------------------------
# FIGURE 2
#---------------------------------------------------------------------------------------------------------------------------
class figure2_v4:
    #--------------------------------------------------------------
    # BASIC FUNCTIONS
    def __init__( self, emdat, pano, width_figure=21 ):
        self.emdat = emdat
        self.pano = pano
        self.width_figure = width_figure
        self.fontsize = 20
        self.fontsizes = {'pano_legend':0.925*self.fontsize, 'pano_legend_title':1.0*self.fontsize, 'pano_axis_ticks':0.85*self.fontsize,\
                          'pano_axis_labels':0.90*self.fontsize, 'pano_title':0.90*self.fontsize, 'pano_txt':0.8*self.fontsize }
        self.dico_lw = {'rect_cases':3, 'global_map':3, 'arrow':2, \
                        'reg_map_country':2, 'reg_map_reg':3, \
                        'gmt_central':3, 'gmt_uncertainties':1, 'gmt_event':3, 'gmt_median_ip':5,\
                        'return_wwo':3, 'return_lines':3, 'return_arrows':3}
        self.arrow_style = ArrowStyle('Fancy', head_length=0.5, head_width=0.5, tail_width=0.25)
        self.dico_cols = {'with_CC':CB_color_cycle[0], 'without_CC':CB_color_cycle[8], 'event':'k', 'counter_event':'grey', 'attrib':'r', 'uncertainties':'grey'}
        
    def plot( self, path_save):
        self.fig = plt.figure( figsize=(self.width_figure, 0.5 * self.width_figure) )
        self.spec = gridspec.GridSpec(ncols=3, nrows=1, figure=self.fig, width_ratios=[1,1,1], height_ratios=[1], left=0.05, right=0.975, bottom=0.125, top=0.970, wspace=0.20, hspace=0.05 )

        # adding panorama
        values_main = xr.where( self.pano.contribs_all['values_global_PR_median'] < 1.e4, self.pano.contribs_all['values_global_PR_median'], 1.e4 )
        values_suppl = self.pano.contribs_all['values_global_I_median']

        list_subax_pano = []
        for i_period, period in enumerate( [[2000,2009], [2010,2019], [2020,2022]] ):
            # selecting only the events that are within the period
            ind_evts = self.emdat['DisNo.'].where( (self.emdat['Start Year'] >= period[0]) & (self.emdat['Start Year'] <= period[1]), drop=True )

            # ploting pano over period
            subax_pano = self.plot_pano(spec_c=self.spec[0,i_period], period=period, i_period=i_period, \
                                        values_main=values_main.sel(event=ind_evts), bounds_main=[1.e0, 1.e1, 1.e2, 1.e3, 1.e4, np.inf],\
                                        values_suppl=values_suppl.sel(event=ind_evts), bounds_suppl=[0.25, 1.0, 2.0, 3.5],\
                                        units={'main':'', 'suppl':u'\u00B0C'}, list_col_suppl = ['gold', 'darkorange', 'red', 'darkred'], integer_legend=False)
            list_subax_pano.append( subax_pano )

        # same y-lim for all
        ylim = [subax_pano.get_ylim() for subax_pano in list_subax_pano]
        ymin, ymax = 0, np.max(np.array(ylim)) # lowest value at 0 anyway
        ticks = np.array(np.arange(ymin, ymax+1, 10), dtype=int) # 10*(ymax//10)+0.1
        for subax_pano in list_subax_pano:
            subax_pano.set_ylim( ymin, ymax )
            subax_pano.set_yticks( ticks=ticks, labels=ticks, size=self.fontsizes['pano_axis_ticks'] )
            
        # save
        self.fig.savefig( os.path.join( path_save, 'figure2_v4.png' ), dpi=300 )
        self.fig.savefig( os.path.join( path_save, 'figure2_v4.pdf' ) )
        return self.fig
    #--------------------------------------------------------------

    @staticmethod
    def sci_notation(number, sig_fig=2, opt_integer=True):
        if np.isnan(number) or np.isinf(number):
            return "$+\infty$"
        else:
            ret_string = "{0:.{1:d}e}".format(number, sig_fig)
            a, b = ret_string.split("e")
            # remove leading "+" and strip leading zeros
            b = int(b)
            if b <= sig_fig:
                if opt_integer and (int(float(a)) == float(a)):
                    return "$" + str(int(float(a)*10**b)) + "$"
                else:
                    return "$" + str(float(a)*10**b) + "$"
            else:
                if opt_integer and (int(float(a)) == float(a)):
                    return "$" + str(int(float(a))) + ".10^{" + str(b) + "}$"
                else:
                    return "$" + a + ".10^{" + str(b) + "}$"

    
    #--------------------------------------------------------------
    # PLOT PANORAMA
    def plot_pano( self, spec_c, period, i_period, values_main, bounds_main, values_suppl, bounds_suppl, units, list_col_suppl, integer_legend ):
        ax_window = plt.subplot( spec_c )
        n_evts_period = values_main.size

        # preparing bins for main axis
        x_main = []
        for i in np.arange(len(bounds_main)-1):
            if np.isinf(bounds_main[i+1]):
                x_main.append( str(self.sci_notation(bounds_main[i])) + '+ '+units['main'] )
            else:
                x_main.append( str(self.sci_notation(bounds_main[i])) + ' to ' + str(self.sci_notation(bounds_main[i+1])) + ' ' + units['main'] )
                
        # preparing bins for supplementary axis
        x_suppl = []
        for j in np.arange(len(bounds_suppl)-1):
            if np.isinf(bounds_suppl[j+1]):
                x_suppl.append( str(self.sci_notation(bounds_suppl[j], opt_integer=integer_legend)) + '+ '+units['suppl'] )
            else:
                x_suppl.append( str(self.sci_notation(bounds_suppl[j], opt_integer=integer_legend)) + ' to ' + str(self.sci_notation(bounds_suppl[j+1], opt_integer=integer_legend)) + ' ' + units['suppl'] )
        
        # preparing values
        weight_counts, count = {}, 0
        for j, s in enumerate(x_suppl):
            tmp = np.zeros(len(x_main))
            for i, m in enumerate(x_main):
                inds = np.where( (bounds_main[i] <= values_main.values) & (values_main.values < bounds_main[i+1]) & (bounds_suppl[j] <= values_suppl.values) & (values_suppl.values < bounds_suppl[j+1]) )[0]
                tmp[i] = len(inds)
            weight_counts[s] = 100 * tmp / n_evts_period # number of events into %
            count += sum(tmp)
        if count != len(values_main.values):
            raise Exception('Bins dont cover all values, enlarge the bounds.')
        
        # preparing size of subplots before text
        xl = [-0.5, len(x_main)-0.5] # doesnt need additional column
        ax_window.set_xlim( xl )
        higher_xmain_categ = np.max( [np.sum([weight_counts[t][i] for t in x_suppl]) for i in np.arange(len(bounds_main)-1)] )
        #yl = [0, (higher_xmain_categ // 10**int(np.log10(higher_xmain_categ)) + 1) * 10**int(np.log10(higher_xmain_categ))] # bound to upper value: eg 64 --> 70; 114 --> 200; etc. Not optimal but works for now.
        yl = [0, 1.05 * higher_xmain_categ]
        ax_window.set_ylim( yl )
        
        # plot
        bottom = np.zeros(len(x_main))
        for i_suppl, suppl in enumerate(weight_counts.keys()):
            _ = ax_window.bar(x=np.arange(len(x_main)), height=weight_counts[suppl], width=0.7, bottom=bottom, label=suppl, facecolor=list_col_suppl[i_suppl] )
            # adding text for each
            for i, x in enumerate(weight_counts[suppl]):
                if x > 0:
                    val = int(np.round(x, 0))
                    t = ax_window.text(x=i, y=bottom[i]+0.5*x, s=str(val)+'%', fontdict={'size':self.fontsizes['pano_txt']},\
                                       verticalalignment='center', horizontalalignment='center', rotation_mode='anchor', bbox={'facecolor':'white', 'alpha':0, 'edgecolor':'white'}, zorder=100)
                    width, pad, shift_y = 0.3, 4, -0.020*(yl[1]-yl[0])#0.3, 1.25, -0.030*(yl[1]-yl[0])
                    height = ((t.get_size()+2*pad)/72.) / (ax_window.get_window_extent().height/self.fig.dpi)
                    #rect = plt.Rectangle( ((i-width/2.),(bottom[i]+0.5*x+shift_y-height/2.)/(yl[1]-yl[0])), width=width, height=height,\
                    #                     transform=ax_window.get_xaxis_transform(), zorder=3, alpha=0.5, fill=True, facecolor='white', edgecolor=list_col_suppl[i_suppl], clip_on=False)
                    #rect = FancyBboxPatch( ((i-width/2.), (bottom[i]+0.5*x+shift_y-height/2.)/(yl[1]-yl[0])), width=width, height=height, boxstyle='Round, pad=1, rounding_size=0.05',\
                    #                      transform=ax_window.get_xaxis_transform(), zorder=3, alpha=0.75, fill=True, facecolor='white', edgecolor=None, lw=0, clip_on=False)
                    #ax_window.add_patch(rect)
            bottom += weight_counts[suppl]

        # polishing
        ax_window.set_xlabel( 'Probability ratio (-)', size=self.fontsizes['pano_axis_labels'] )
        ax_window.set_ylabel( 'Allocation of events in each category (%)', size=self.fontsizes['pano_axis_labels'] )
        ax_window.set_xticks( ticks=np.arange(len(x_main)), labels=x_main, size=self.fontsizes['pano_axis_ticks'], rotation=-15 )
        #ax_window.set_yticks( ticks=ax_window.get_yticks(), labels=['' for s in ax_window.get_yticks()], size=self.fontsizes['pano_axis_ticks'] )
        ax_window.set_title('('+list_letters_panels[i_period] + "): "+str(n_evts_period)+" heatwaves attributed over "+str(period[0])+"-"+str(period[1]), size=self.fontsizes['pano_title'], fontweight='bold')
        handles, labels = ax_window.get_legend_handles_labels()
        legend = ax_window.legend(handles[::-1], labels[::-1], frameon=True, markerfirst=False, loc=0,\
                                  title="Change in intensity", prop={'size':self.fontsizes['pano_legend']}, title_fontsize=self.fontsizes['pano_legend_title'])
        return ax_window
    #--------------------------------------------------------------
#---------------------------------------------------------------------------------------------------------------------------
#---------------------------------------------------------------------------------------------------------------------------





















#---------------------------------------------------------------------------------------------------------------------------
# FIGURE 3
#---------------------------------------------------------------------------------------------------------------------------
class figure3_v4:
    #--------------------------------------------------------------
    # BASIC FUNCTIONS
    def __init__( self, emissions_FF, emissions_GCB, dGMT_OSCAR, data_to_do, emissions_Jones023, pano, entities_plot, emdat, method_entity, width_figure=21, fraction_tier1=0.5, level_OSCAR='mean' ):
        self.emissions_FF = emissions_FF
        self.emissions_GCB = emissions_GCB
        self.dGMT_OSCAR = dGMT_OSCAR
        self.data_to_do = data_to_do
        self.emissions_Jones023 = emissions_Jones023
        self.emdat = emdat
        self.pano = pano
        self.entities_plot = entities_plot
        self.method_entity = method_entity
        self.width_figure = width_figure
        self.fraction_tier1 = fraction_tier1
        self.level_OSCAR = level_OSCAR
        self.yr_arrow = 2019
        self.x_arrow, self.h_arrow1 = 0.02, 0.0075
        self.y_arrow0, self.y_arrow1, self.yarrow_others = 0.88, 0.050, 0.45
        self.years_plot = np.arange(1950, 2022+1)
        self.years_integration = np.arange(1850,2022+1)
        self.year_ref = 2022
        self.space_time_bar, self.width_time_bar = 0.005, 0.03
        self.lw_thin_entities = 0.125

        # other values
        self.year_GMT = 2022
        self.fontsize = 20
        self.fontsizes = {'ticks':0.8*self.fontsize, 'entity':0.75*self.fontsize, 'legend':0.9*self.fontsize, 'label':0.8*self.fontsize, 'bar_txt':0.8*self.fontsize, 'title':0.9*self.fontsize,\
                          'GMT_labels':0.9*self.fontsize, 'GMT_ticks':0.8*self.fontsize, 'GMT_text':1.0*self.fontsize, \
                          'pano_title':1.0*self.fontsize, 'pano_labels':0.85*self.fontsize, 'pano_axis_labels':0.85*self.fontsize,\
                          'pano_axis_ticks':0.7*self.fontsize, 'pano_ticks':0.7*self.fontsize, 'pano_axis_ticks':0.7*self.fontsize, 'pano_txt':0.7*self.fontsize, 'pano_legend':0.8*self.fontsize, 'pano_legend_title':0.85*self.fontsize, \
                          'group_title':1.0*self.fontsize
                         }
        
        self.arrow_style = ArrowStyle('Fancy', head_length=0.5, head_width=0.5, tail_width=0.25)
        self.dico_cols_bars = {'GCB':'red', 'Jones2023':'red', 'ERA5':'red', 'BEST':'purple',\
                               'Remaining':'lightgrey', 'Other majors':'grey', 'separation majors':'k', 'Biggest majors':'saddlebrown'}        
        self.dico_lw = {'rect_entities':2, 'entities_tier2':0.25}
        self.vmax_pano_pct = 33 #%
        self.alpha_subplots = 0.75
        self.append_ghost_left = False
        self.fact_pad_bbox = {'left':0.18, 'right':0.08, 'low':0.78, 'up':0.14}
        self.width_ratios = [1,1,1]
        self.height_ratios = [1,1, 1,1,1]
        self.rough_dico_spec_entities = {0:[2,0], 1:[2,1], 2:[2,2], 3:[3,0], 4:[3,1], 5:[3,2], 6:[4,0], 7:[4,1], 8:[4,2]}
    
    def plot( self, path_save, arrow_or_legend='legend' ):
        # define entities to plot
        self.define_entities_plot()
        self.local_dico_col_entities,count_nogray = {},-1
        for i_entity, entity in enumerate(self.list_entities_plot):
            if i_entity % 7 == 0:
                count_nogray += 1
            self.local_dico_col_entities[entity] = tuple(np.array(CB_color_cycle[i_entity+count_nogray]) * 0.8) # darkening colors to facilitate readibility
        
        # creating figure
        self.fig = plt.figure( figsize=(self.width_figure, 1.0*self.width_figure) )
        self.fig_renderer = self.fig.canvas.get_renderer()
        self.spec = gridspec.GridSpec(ncols=len(self.width_ratios), nrows=len(self.height_ratios), figure=self.fig, height_ratios=self.height_ratios, width_ratios=self.width_ratios, left=0.10, right=0.975, bottom=0.10, top=0.95, wspace=0.50, hspace=0.60 )

        # preparation of entities and categories
        self.entities_tier1_CO2, self.entities_tier2_CO2  = eval_tiers_emissionsCO2(emissions_FF=self.emissions_FF, fraction_tier1=self.fraction_tier1, years_integration=self.years_integration)
        self.entities_tier1_GMT, self.entities_tier2_GMT = eval_tiers_dGMT(dGMT_OSCAR=self.dGMT_OSCAR, fraction_tier1=self.fraction_tier1, years_integration=self.years_integration, level_OSCAR=self.level_OSCAR)
        self.factor_unit_FFCO2, self.factor_unit_GCB, self.unit_CO2 = 1.e-3, 1.e-3, 'GtCO$_2$'
        self.factor_unit_FFCH4, self.factor_unit_Jones2023, self.unit_CH4 = 1., 1., 'MtCH$_4$' # equiv TgCH4 yr-1
        self.add_LUC_GCB = True
        
        # CO2
        ax = plt.subplot(self.spec[0:1+1,0])
        ax = self.plot_timeseries( ax=ax, variable='cumulative_emissions_CO2', arrow_or_legend=arrow_or_legend, do_2ndlegend=False, letter_panel=list_letters_panels[0])
        pos_new = ax.get_position()
        ax_bis = self.fig.add_axes([pos_new.x1+self.space_time_bar, pos_new.y0, self.width_time_bar, pos_new.height])
        ax_bis = self.plot_bar( ax=ax_bis, variable='cumulative_emissions_CO2' )

        # temperatures
        ax = plt.subplot(self.spec[0:1+1,1])
        ax = self.plot_timeseries( ax=ax, variable='temperatures', arrow_or_legend=arrow_or_legend, do_2ndlegend=True, letter_panel=list_letters_panels[1])
        pos_new = ax.get_position()
        ax_bis = self.fig.add_axes([pos_new.x1+self.space_time_bar, pos_new.y0, self.width_time_bar, pos_new.height])
        ax_bis = self.plot_bar( ax=ax_bis, variable='temperatures' )

        # preparing panorama
        values_main = xr.where( self.pano.contribs_all['values_'+self.method_entity+'_PR_median'] < 1.e4, 1 + self.pano.contribs_all['values_'+self.method_entity+'_PR_median'], 1.e4 )
        values_suppl = self.pano.contribs_all['values_'+self.method_entity+'_I_median']
        
        # ploting each entity
        dico_pos_ax, list_subax_pano = {}, []
        for i_entity, entity in enumerate(self.list_entities_plot):
            row_spec, col_spec = self.rough_dico_spec_entities[i_entity]
            ax = plt.subplot(self.spec[row_spec, col_spec])
            if False:
                ax = self.plot_pano_a(ax=ax, entity=entity, values_main=values_main, values_suppl=values_suppl, units={'main':'', 'suppl':u'\u00B0C'}, \
                                    bounds_main=[1.0, 1.1, 10, 100, 1000, np.inf], ticks_main=[1, 1.1, 10, 100, 1000, 10000],\
                                    bounds_suppl=[0., 0.01, 0.05, 0.10, 0.20, 0.40], ticks_suppl=[0., 0.01, 0.05, 0.10, 0.20, 0.40])
            else:
                ax = self.plot_pano_b(ax_window=ax, entity=entity, values_main=values_main, values_suppl=values_suppl, units={'main':'', 'suppl':u'\u00B0C'}, \
                                    bounds_main=[1.0, 1.1, 10, 100, 1000, np.inf], bounds_suppl=[0., 0.01, 0.05, 0.10, 0.20, 0.40],\
                                    list_col_suppl = ['yellow', 'gold', 'darkorange', 'red', 'darkred'], integer_legend=False, do_legend=(i_entity==7))
            # labels_main = ['\n$< +10\%$', '$+10\%$\nto\n$+100\%$', '$+100\%$\nto\n$1.10^4$', '\n$> 1.10^4$' ]
            #list_letters_panels[0]
            dico_pos_ax[i_entity] = ax.get_position()
            list_subax_pano.append( ax )

        # same y-lim for all
        ylim = [ax.get_ylim() for ax in list_subax_pano]
        ymin, ymax = 0, np.max(np.array(ylim)) # lowest value at 0 anyway
        ticks = np.array(np.arange(ymin, ymax+1, 10), dtype=int) # 10*(ymax//10)+0.1
        for ax in list_subax_pano:
            ax.set_ylim( ymin, ymax )
            ax.set_yticks( ticks=ticks, labels=ticks, size=self.fontsizes['pano_axis_ticks'] )

        # box around this group
        x_left = dico_pos_ax[0].x0 - self.fact_pad_bbox['left'] * dico_pos_ax[0].width
        x_right = dico_pos_ax[2].x1 + self.fact_pad_bbox['right'] * dico_pos_ax[0].width
        y_top = dico_pos_ax[0].y1 + self.fact_pad_bbox['up'] * dico_pos_ax[0].height
        y_bottom = dico_pos_ax[i_entity].y0 - self.fact_pad_bbox['low'] * dico_pos_ax[0].height
        rect = plt.Rectangle((x_left, y_bottom), x_right-x_left, y_top - y_bottom,\
                             edgecolor='k', facecolor=(0,0,0,0), linewidth=self.dico_lw['rect_entities'], zorder=-1000)
        self.fig.add_artist(rect)
        n_evt = pano.contribs_all.event.size
        period = str(int(emdat['Start Year'].min())) + '-'+ str(int(emdat['End Year'].max()))
        t = self.fig.text(x_left+0.5*(x_right-x_left), y_top+0.0075, 'Attribution of '+str(n_evt)+' heatwaves over '+period+' to the carbon majors: allocation of the events based on their contributions',\
                          fontdict={'size':self.fontsizes['group_title'], 'weight':'bold'}, verticalalignment='center', horizontalalignment='center', rotation_mode='anchor', bbox={'facecolor':'white', 'alpha':0, 'edgecolor':'white'})
        
        # save
        self.fig.savefig( os.path.join( path_save, 'figure3-'+self.method_entity+'_v4.png' ), dpi=300 )
        self.fig.savefig( os.path.join( path_save, 'figure3-'+self.method_entity+'_v4.pdf' ) )
        return self.fig
    #--------------------------------------------------------------

    

    
    #--------------------------------------------------------------
    # NEW FUNCTIONS FROM FIGURE 2
    def get_data(self, variable):
        # preparation data
        if variable in ['emissions_CO2', 'cumulative_emissions_CO2', 'emissions_CH4', 'cumulative_emissions_CH4']:
            # CO2 reference
            if variable in ['emissions_CO2', 'cumulative_emissions_CO2']:
                if self.add_LUC_GCB:
                    values_tot = (self.emissions_GCB['FF_CO2'] + self.emissions_GCB['cement_CO2'] + self.emissions_GCB['LUC_CO2']).sel(year=self.years_integration) * self.factor_unit_GCB
                else:
                    values_tot = (self.emissions_GCB['FF_CO2'] + self.emissions_GCB['cement_CO2']).sel(year=self.years_integration) * self.factor_unit_GCB

            # CH4 reference
            elif variable in ['emissions_CH4', 'cumulative_emissions_CH4']:
                if self.add_LUC_GCB:
                    values_tot = (self.emissions_Jones023['emissions_CH4_Fossil'] + self.emissions_Jones023['emissions_CH4_LULUCF']).sel(year=self.years_integration) * self.factor_unit_Jones2023
                else:
                    values_tot = self.emissions_Jones023['emissions_CH4_Fossil'].sel(year=self.years_integration) * self.factor_unit_Jones2023

            # cumulative?
            if variable in ['cumulative_emissions_CO2', 'cumulative_emissions_CH4']:
                values_tot = values_tot.cumsum( 'year' )
            values_tot = values_tot.sel(year=self.years_plot)

            # data entities
            if variable in ['emissions_CO2', 'cumulative_emissions_CO2']:
                values_entities = self.emissions_FF['emissions_CO2'].sel(year=self.years_integration) * self.factor_unit_FFCO2
            elif variable in ['emissions_CH4', 'cumulative_emissions_CH4']:
                values_entities = self.emissions_FF['emissions_CH4'].sel(year=self.years_integration) * self.factor_unit_FFCH4
            if variable in ['cumulative_emissions_CO2', 'cumulative_emissions_CH4']:
                values_entities = values_entities.cumsum('year')
            values_entities = values_entities.sel(year=self.years_plot)

            # taking same order for CO2 and CH4, simplification for plots
            #entities_tier1, entities_tier2 = self.entities_tier1_CO2, self.entities_tier2_CO2
            
        elif variable == 'temperatures':
            values_tot = {}
            # plot data reference
            for obs in ['ERA5', 'BEST']:
                values_tot[obs] = self.data_to_do[obs][1] - (self.data_to_do[obs][1].sel(time=slice(1961,1990)).mean() - 0.36) # IPCC AR6 Chapter2, Cross Chapter Box 2.3, Table 1

            # order
            #entities_tier1, entities_tier2 = self.entities_tier1_GMT, self.entities_tier2_GMT

            # data entities
            values_entities = self.dGMT_OSCAR['dGMT_entities_mean'].sel(time=self.years_plot)

        else:
            raise Exception('Wrong name of variable, check prepared ones.')
            
        # because different composition of tiers depending on CO2 and GMT, choosing one as reference to have same
        if self.method_entity == 'GMT':
            entities_tier1, entities_tier2 = self.entities_tier1_GMT, self.entities_tier2_GMT
        else:
            entities_tier1, entities_tier2 = self.entities_tier1_CO2, self.entities_tier2_CO2
        
        return values_tot, values_entities, entities_tier1, entities_tier2

    
    def plot_timeseries( self, ax, variable, arrow_or_legend, letter_panel, do_2ndlegend=False ):
        # preparation data
        values_tot, values_entities, entities_tier1, entities_tier2 = self.get_data(variable)
        plot_1st_legend, plot_2nd_legend = [],  []
        if variable in ['emissions_CO2', 'cumulative_emissions_CO2', 'emissions_CH4', 'cumulative_emissions_CH4']:
            # plot data reference
            if variable in ['emissions_CO2', 'cumulative_emissions_CO2']:
                pp = plt.plot( self.years_plot, values_tot, color=self.dico_cols_bars['GCB'], lw=4, label='GCB 2023' )
            elif variable in ['emissions_CH4', 'cumulative_emissions_CH4']:
                pp = plt.plot( self.years_plot, values_tot, color=self.dico_cols_bars['Jones2023'], lw=4, label='Jones et al., 2023' )
            plot_1st_legend.append( pp[0] )

            # values to iterate fill_between
            tmp_bottom = values_tot.values
            
            # additional properties
            self.ylim = 0, 1.05*np.nanmax(values_tot.values)
            
        elif variable == 'temperatures':
            # plot data reference
            for obs in ['ERA5', 'BEST']:
                pp = plt.plot( values_tot[obs].time, values_tot[obs].values, color=self.dico_cols_bars[obs], lw=4, ls={'ERA5':'-', 'BEST':'--'}[obs], label=obs )
                plot_1st_legend.append( pp[0] )
                
            # values to iterate fill_between
            tmp_bottom = values_tot['ERA5'].sel(time=self.years_plot).values

            # additional properties
            #self.ylim = np.min(tmp_bottom) - 0.10*(np.max(tmp_bottom)-np.min(tmp_bottom)), np.max(tmp_bottom) + 0.05*(np.max(tmp_bottom)-np.min(tmp_bottom))
            self.ylim = 0, 1.05*np.nanmax(tmp_bottom)
        
        # preparing rest of plot
        xlim = self.years_plot[0], self.years_plot[-1]
        ax.set_xlim(xlim)
        ax.set_ylim(self.ylim)

        # adapt order of entities if emissions, to avoid croissing when entity disappears
        tmp_entities = list(np.copy(entities_tier1))
        if variable in ['emissions_CO2', 'emissions_CH4']:
            def test_len(entity):
                return np.where(~np.isnan(self.emissions_FF['emissions_CO2'].sel(entity=entity, year=self.years_plot).values))[0][-1]
            tmp_entities.sort(key=test_len, reverse=True)        
        
        # Tier 1
        for j, entity in enumerate(tmp_entities):
            # selecting values
            values_entity = values_entities.sel(entity=entity).values
            
            # dealing with missing emissions for nice representations (no effect on temperatures)
            ind_last = np.where(~np.isnan(values_entity))[0][-1]
            yr_last = np.min( [self.yr_arrow, self.years_plot[ind_last]] )
            # to avoid having a blank triangle with next entity
            values_entity[np.where(np.isnan(values_entity))] = 0
            
            # plot
            if entity in dict_shorten_entities:
                lbl = dict_shorten_entities[entity]
            else:
                lbl = entity
            pp = plt.fill_between( self.years_plot, tmp_bottom, tmp_bottom-values_entity, alpha=0.85,\
                                  color=self.local_dico_col_entities[entity], lw=self.lw_thin_entities, edgecolor=self.dico_cols_bars['separation majors'], label=lbl )
            plot_2nd_legend.append( pp )
            if arrow_or_legend == 'arrow':
                # adding name and arrow
                self.do_arrow_text(x_start=xlim[0]+self.x_arrow*(xlim[1]-xlim[0]), y_start=self.ylim[0] + (self.y_arrow0 - j*self.y_arrow1) * (self.ylim[1]-self.ylim[0]),\
                                   x_end=yr_last, y_end=(tmp_bottom-0.5*values_entity)[yr_last-self.years_plot[0]],\
                                   text=lbl, fontcolor_arrow=self.local_dico_col_entities[entity], ax=ax)
            tmp_bottom[np.where(~np.isnan(values_entity))] -= values_entity[np.where(~np.isnan(values_entity))]
        
        # Tier 2: at bottom
        mem = np.copy(tmp_bottom)
        for i, entity in enumerate(entities_tier2):
            # selecting values
            values_entity = values_entities.sel(entity=entity).values

            # dealing with missing emissions for nice representations (no effect on temperatures)
            if np.all(np.isnan(values_entity)):
                values_entity = np.zeros( values_entity.shape )
            else:
                ind_last = np.where(~np.isnan(values_entity))[0][-1]
                yr_last = np.min( [self.yr_arrow, self.years_plot[ind_last]] )
                # to avoid having a blank triangle with next entity
                values_entity[np.where(np.isnan(values_entity))] = 0
            
            # plot
            if entity in dict_shorten_entities:
                lbl = dict_shorten_entities[entity]
            else:
                lbl = entity
            if entity in self.list_entities_plot:
                pp = plt.fill_between(self.years_plot, tmp_bottom, tmp_bottom-values_entity, alpha=0.85,\
                                      color=self.local_dico_col_entities[entity], lw=self.lw_thin_entities, edgecolor=self.dico_cols_bars['separation majors'], label=lbl )
                plot_2nd_legend.append( pp )
            else:
                pp = plt.fill_between(self.years_plot, tmp_bottom, tmp_bottom-values_entity, alpha=0.85,\
                                      color=self.dico_cols_bars['Other majors'], lw=self.lw_thin_entities, edgecolor=self.dico_cols_bars['separation majors'], label=str(emissions_FF.entity.size-len(self.list_entities_plot))+' other carbon majors' )
            tmp_bottom[np.where(~np.isnan(values_entity))] -= values_entity[np.where(~np.isnan(values_entity))]
        plot_2nd_legend.append( pp ) # add only once the last one
        if arrow_or_legend == 'arrow':
            # adding global name
            self.do_arrow_text(x_start=xlim[0]+self.x_arrow*(xlim[1]-xlim[0]), y_start=self.ylim[0] + self.yarrow_others * (self.ylim[1]-self.ylim[0]),\
                               x_end=self.yr_arrow, y_end=(tmp_bottom + 0.25*(mem-tmp_bottom))[self.yr_arrow-self.years_plot[0]],\
                               text='Other carbon\nmajors', fontcolor_arrow=self.dico_cols_bars['Other majors'], ax=ax)
        
        # remaining category
        pp = plt.fill_between( self.years_plot, tmp_bottom, np.zeros(len(tmp_bottom)), alpha=0.85, color=self.dico_cols_bars['Remaining'], edgecolor=self.dico_cols_bars['Remaining'], label='Other actors' )
        plot_2nd_legend.append( pp )
        if arrow_or_legend == 'arrow':
            # adding name
            if variable in ['emissions_CO2', 'cumulative_emissions_CO2']:
                txt = {True:'Remaining CO$_2$ emissions', False:'Remaining CO$_2$ emissions from\nfossil fuels & cement'}[self.add_LUC_GCB]
            elif variable in ['emissions_CH4', 'cumulative_emissions_CH4']:
                txt = {True:'Remaining CH$_4$ emissions', False:'Remaining CH$_4$ emissions from\nfossil fuels & cement'}[self.add_LUC_GCB]
            else:
                txt = 'Remaining climate change'
            self.do_arrow_text(x_start=xlim[0]+0.4*(xlim[1]-xlim[0]), y_start=0.025*(self.ylim[1]-self.ylim[0]),\
                               with_arrow=False, x_end=None, y_end=None,\
                               text=txt, fontcolor_arrow=None, ax=ax)#0.25*tmp_bottom[-1]

        # ticks
        _ = plt.xticks( size=self.fontsizes['ticks'], rotation=-20 )
        _ = plt.yticks( size=self.fontsizes['ticks'] )
        
        # y-label
        if variable in ['emissions_CO2', 'cumulative_emissions_CO2', 'emissions_CH4', 'cumulative_emissions_CH4']:
            txt = {'emissions_CO2':'CO$_2$ emissions',\
                   'cumulative_emissions_CO2':'Cumulative CO$_2$ emissions',\
                   'emissions_CH4':'CH$_4$ emissions',\
                   'cumulative_emissions_CH4':'Cumulative CH$_4$ emissions'}[variable]
            txt_luc = {True:'\nfrom all sectors', False:'\n from fossil fuels & cement'}[self.add_LUC_GCB]
            txt_cum = {'emissions_CO2':'',\
                       'cumulative_emissions_CO2':' since '+str(self.years_integration[0]),\
                       'emissions_CH4':'',\
                       'cumulative_emissions_CH4':' since '+str(self.years_integration[0])}[variable]
            txt_unit = {'emissions_CO2':self.unit_CO2,\
                        'cumulative_emissions_CO2':self.unit_CO2,\
                        'emissions_CH4':self.unit_CH4,\
                        'cumulative_emissions_CH4':self.unit_CH4}[variable]
            plt.title( '('+letter_panel+'): '+txt + txt_luc + txt_cum + ' ('+txt_unit+')', size=self.fontsizes['title'] )
        else:
            plt.title( '('+letter_panel+'): '+'Change in Global Mean Temperature\nwith reference to 1850-1900 ('+ u'\u00B0C'+')', size=self.fontsizes['title'] )
        
        # 1st legend
        lgd1 = plt.legend( plot_1st_legend, [pp.get_label() for pp in plot_1st_legend], loc='upper left', prop={'size':self.fontsizes['legend']}, frameon=False, markerfirst=False )
        ax.add_artist(lgd1)
        
        # 2nd legend
        if do_2ndlegend:
            #lgd2 = plt.legend( plot_2nd_legend, [pp.get_label() for pp in plot_2nd_legend], loc='center', bbox_to_anchor=(0.1,-0.07), ncol=4, prop={'size':self.fontsizes['legend']}, frameon=True, markerfirst=False )
            lgd2 = plt.legend( plot_2nd_legend, [pp.get_label() for pp in plot_2nd_legend], loc='center', bbox_to_anchor=(1.9,0.5), ncol=1, prop={'size':self.fontsizes['legend']}, frameon=True, markerfirst=True )
            ax.add_artist(lgd2)
        return ax

    
    def do_arrow_text( self, x_start, y_start, x_end, y_end, text, fontcolor_arrow, ax, with_arrow=True):
        xl = ax.get_xlim()
        yl = ax.get_ylim()
        t = ax.text((x_start-xl[0]) / (xl[1]-xl[0]), (y_start-yl[0]) / (yl[1]-yl[0]), text, transform=ax.transAxes, fontdict={'size':self.fontsizes['entity']},\
                    verticalalignment='center', horizontalalignment='left', rotation_mode='anchor', bbox={'facecolor':'white', 'alpha':0, 'edgecolor':'white'})
        if with_arrow:
            t_box = t.get_window_extent(renderer=self.fig_renderer).transformed(ax.transData.inverted())
            ax.add_patch( FancyArrowPatch((t_box.x1 + 0.01*(xl[1]-xl[0]), t_box.y0 + 0.5*t_box.height), (x_end, y_end), \
                                          mutation_scale=25, arrowstyle=self.arrow_style, facecolor=fontcolor_arrow, edgecolor='k', zorder=1000) )
            
    
    def plot_bar(self, ax, variable):
        # preparation data
        values_tot, values_entities, entities_tier1, entities_tier2 = self.get_data(variable)
        if variable in ['emissions_CO2', 'cumulative_emissions_CO2', 'emissions_CH4', 'cumulative_emissions_CH4']:
            val_tot = values_tot.sel(year=self.year_ref).values
            values_entities = values_entities.sel(year=self.year_ref)
        elif variable == 'temperatures':
            val_tot = values_tot['ERA5'].sel(time=self.year_ref).values
            values_entities = values_entities.sel(time=self.year_ref)
        
        # Calculating contributions
        val_tier1 = values_entities.sel(entity=entities_tier1).sum()
        val_tier2 = values_entities.sel(entity=entities_tier2).sum()
        values = {'Remaining': np.array([((val_tot-val_tier1-val_tier2)/val_tot).values]) }
        values['Other majors'] = np.array([(val_tier2/val_tot).values])
        for entity in entities_tier2[::-1]:
            values[entity] = np.array([(values_entities.sel(entity=entity)/val_tot).values])
        values['Biggest majors'] = np.array([(val_tier1/val_tot).values])
        for entity in entities_tier1[::-1]:
            values[entity] = np.array([(values_entities.sel(entity=entity)/val_tot).values])
        
        # bar plot, with an overall bar for all entities tier 1
        ax.set_ylim(self.ylim[0], self.ylim[1])
        bottom, width_bar, frac_biggest = np.array([0.]), 0.8, 0.25#0.125
        for val in values:
            if val in ['Remaining', 'Other majors']:
                _ = ax.bar(x=0, height=values[val] * val_tot, width=width_bar, bottom=bottom, label=val, facecolor=self.dico_cols_bars[val] )
            elif val in entities_tier1:
                _ = ax.bar(x=-0.5*frac_biggest*width_bar, height=values[val] * val_tot, width=(1-frac_biggest)*width_bar, bottom=bottom, label=val, facecolor=self.local_dico_col_entities[val] )
            elif val in entities_tier2:
                if val in self.list_entities_plot:
                    col = self.local_dico_col_entities[val]
                else:
                    col = self.dico_cols_bars['Other majors']
                _ = ax.bar(x=-0.5*frac_biggest*width_bar, height=values[val] * val_tot, width=(1-frac_biggest)*width_bar, bottom=bottom,\
                           label=val, facecolor=col )
            elif val in ['Other majors']: # other represents a fraction of the complete bar
                _ = ax.bar(x=0.5*(width_bar-frac_biggest*width_bar), height=values[val] * val_tot, width=frac_biggest*width_bar, bottom=bottom,\
                           label=val, facecolor=self.dico_cols_bars['separation majors'] )
            elif val in ['Biggest majors']: # biggest represents a fraction of the complete bar
                _ = ax.bar(x=0.5*(width_bar-frac_biggest*width_bar), height=values[val] * val_tot, width=frac_biggest*width_bar, bottom=bottom,\
                           label=val, facecolor=self.dico_cols_bars[val] )
            ax.hlines(y=bottom+values[val]*val_tot, xmin=-0.5*width_bar, xmax=(0.5-frac_biggest)*width_bar, lw=self.lw_thin_entities, color=self.dico_cols_bars['separation majors'])
            
            # text for information
            if True:
                if val in ['Remaining', 'Other majors', 'Biggest majors']:
                    # adding text in box OUTSIDE BAR:
                    t = ax.text(x=1.95*0.5*width_bar, y=bottom + 0.5*values[val] * (self.ylim[1]-self.ylim[0]), s=str(int(np.round(100*values[val]))) + '%', fontdict={'size':self.fontsizes['bar_txt']},\
                                verticalalignment='center', horizontalalignment='center', rotation_mode='anchor', bbox={'facecolor':'white', 'alpha':0, 'edgecolor':'white'}, zorder=100)
            else:
                # adding text in box INSIDE BAR:
                t = ax.text(x=0, y=bottom + 0.5*values[val] * (self.ylim[1]-self.ylim[0]), s=str(int(np.round(100*values[val]))) + '%', fontdict={'size':self.fontsizes['bar_txt']},\
                            verticalalignment='center', horizontalalignment='center', rotation_mode='anchor', bbox={'facecolor':'white', 'alpha':0, 'edgecolor':'white'}, zorder=100)
                width, pad, shift_y = 0.75, 5, 0.175
                height = ((t.get_size()+2*pad)/72.) / (ax.get_window_extent().height/self.fig.dpi)
                rect = FancyBboxPatch( (-0.5*width, (bottom + 0.5*values[val]*val_tot - self.ylim[0])[0] / (self.ylim[1] - self.ylim[0]) + shift_y*(1 - val_tot/(self.ylim[1] - self.ylim[0])) - height/2.), width=width, height=height,\
                                      boxstyle='Round, pad=0, rounding_size=0.05', transform=ax.get_xaxis_transform(), zorder=3, alpha=0.75, fill=True, facecolor='white', edgecolor=None, lw=0, clip_on=False)
                ax.add_patch(rect)

            # preparing next round: increasing bottom for all EXCEPT Biggest majors
            if val not in ['Biggest majors', 'Other majors']:
                bottom[0] += values[val][0] * val_tot
        
        # finishing subplot
        plt.axis('off')
        plt.title( 'Share\n'+str(self.year_ref), size=self.fontsizes['label'], y=0.95 )
    #--------------------------------------------------------------



    #--------------------------------------------------------------
    # NEW FUNCTIONS FROM FIGURE 3    
    def define_entities_plot(self):
        # evaluating categories of entities
        if self.method_entity == 'CO2':
            self.entities_tier1, self.entities_tier2  = eval_tiers_emissionsCO2(emissions_FF=self.emissions_FF, fraction_tier1=self.fraction_tier1, years_integration=self.years_integration)
        elif self.method_entity == 'GMT':
            self.entities_tier1, self.entities_tier2 = eval_tiers_dGMT(dGMT_OSCAR=self.dGMT_OSCAR, fraction_tier1=self.fraction_tier1, years_integration=self.years_integration, level_OSCAR=self.level_OSCAR)
        self.entities_tier1, self.entities_tier2 = list(self.entities_tier1), list(self.entities_tier2)

        # preparing list of entities that will be used: broad categories first
        if (('tier1' in self.entities_plot) or ('Tier1' in self.entities_plot))  and  (('tier2' in self.entities_plot) or ('Tier2' in self.entities_plot)):
            self.list_entities_plot = self.entities_tier1 + self.entities_tier2
        elif ('tier1' in self.entities_plot) or ('Tier1' in self.entities_plot):
            self.list_entities_plot = list(self.entities_tier1)
        elif ('tier2' in self.entities_plot) or ('Tier2' in self.entities_plot):
            self.list_entities_plot = list(self.entities_tier2)
        else:
            self.list_entities_plot = []
            
        # preparing list of entities that will be used: details then
        for ent in self.entities_plot:
            if ent not in ['tier1', 'tier2', 'Tier1', 'Tier2']:
                if (ent in self.entities_tier1) and (('tier1' in self.entities_plot) or ('Tier1' in self.entities_plot)):
                    pass # already in there
                elif (ent in self.entities_tier2) and (('tier2' in self.entities_plot) or ('Tier2' in self.entities_plot)):
                    pass # already in there
                else:
                    self.list_entities_plot.append( ent )

        # preparinng names of panels for each entity
        self.dico_panel_entity = {}
        for i_entity, entity in enumerate(self.list_entities_plot):
            row_spec, col_spec = self.rough_dico_spec_entities[i_entity]
            self.dico_panel_entity[entity] = list_letters_panels[ 2+i_entity ]
                    
    @staticmethod
    def sci_notation(number, sig_fig=4, opt_integer=True):
        if np.isnan(number) or np.isinf(number):
            return "$+\infty$"
        else:
            ret_string = "{0:.{1:d}e}".format(number, sig_fig)
            a, b = ret_string.split("e")
            # remove leading "+" and strip leading zeros
            b = int(b)
            if b <= sig_fig:
                if opt_integer and (int(float(a)) == float(a)):
                    return "$" + str(int(float(a)*10**b)) + "$"
                else:
                    return "$" + str(float(a)*10**b) + "$"
            else:
                if opt_integer and (int(float(a)) == float(a)):
                    return "$" + str(int(float(a))) + ".10^{" + str(b) + "}$"
                else:
                    return "$" + a + ".10^{" + str(b) + "}$"

    def plot_pano_a(self, ax, entity, values_main, bounds_main, values_suppl, bounds_suppl, units, ticks_main, ticks_suppl, integer_legend=False):
        # preparing values
        weight_counts = np.zeros( (len(bounds_suppl), len(ticks_main)) )
        for j in np.arange(len(bounds_suppl)-1):  
            for i in np.arange(len(bounds_main)-1):
                inds = np.where((bounds_main[i] <= values_main.sel(entity=entity).values) & (values_main.sel(entity=entity).values < bounds_main[i+1]) &\
                                (bounds_suppl[j] <= values_suppl.sel(entity=entity).values) & (values_suppl.sel(entity=entity).values < bounds_suppl[j+1]) )[0]
                weight_counts[j,i] = len(inds)
        # thus the percentage
        #weight_counts *= 100/np.sum(weight_counts)

        # plot
        norm = plt.Normalize(vmin=0, vmax=self.vmax_pano_pct)
        cmap = plcol.LinearSegmentedColormap.from_list('', ['white',self.local_dico_col_entities[entity]])
        ax.set_xlim(-0.5, len(ticks_main)-1-0.5)
        ax.set_ylim(-0.5, len(ticks_suppl)-1-0.5)
        _ = ax.pcolormesh(np.arange(len(ticks_main)), np.arange(len(ticks_suppl)), weight_counts, cmap=cmap, norm=norm)

        # creating text in box for each cell
        width, height = 0.5, 1
        for i in np.arange(len(bounds_main)-1):
            for j in np.arange(len(bounds_suppl)-1):
                val = int(np.round(weight_counts[j,i],0))
                if val > 0:
                    t = ax.text(x=i, y=j, s=str(val), fontdict={'size':self.fontsizes['pano_txt']},\
                                       verticalalignment='center', horizontalalignment='center', rotation_mode='anchor', bbox={'facecolor':'white', 'alpha':0, 'edgecolor':'white'}, zorder=100)
                    height_box, shift_y = 0.35, 0.0125
                    rect = FancyBboxPatch((i-0.8*width/2., j + shift_y - height_box/2.), width=0.8*width, height=height_box, boxstyle='Round, pad=0, rounding_size=0.05',\
                                          zorder=3, alpha=0.5, fill=True, facecolor='white', edgecolor=None, lw=0, clip_on=False)
                    ax.add_patch(rect)

        # polishing
        if entity in dict_even_shorter_entities:
            lbl = dict_even_shorter_entities[entity]
        elif entity in dict_shorten_entities:
            lbl = dict_shorten_entities[entity]
        else:
            lbl = entity
        ax.set_title( '(' + self.dico_panel_entity[entity] + '): ' + lbl, size=self.fontsizes['pano_title'], loc='center', x=0.5, color=self.local_dico_col_entities[entity] )#, fontweight='bold'
        #ax.set_xlabel( 'Contribution to the probabilities', size=self.fontsizes['pano_labels'] )
        ax.set_xlabel( 'Multiplication of preindustrial probability\ndue to entity (-)', size=self.fontsizes['pano_labels'] )
        #ax.xaxis.set_label_coords(0.50, -0.16)
        #ax.set_ylabel( 'Contribution to the intensities (' + units['suppl'] + ')', size=self.fontsizes['pano_labels'] )
        ax.set_ylabel( 'Change in intensity\ndue to entity (' + units['suppl'] + ')', size=self.fontsizes['pano_labels'] )
        ax.set_xticks( ticks=np.arange(len(ticks_main))-0.5, labels=ticks_main, size=self.fontsizes['pano_ticks'], rotation=0 )#-20
        ax.set_yticks( ticks=np.arange(len(ticks_suppl))-0.5, labels=ticks_suppl, size=self.fontsizes['pano_ticks'] )
        return ax

    def plot_pano_b( self, ax_window, entity, values_main, bounds_main, values_suppl, bounds_suppl, units, list_col_suppl, integer_legend, do_legend ):
        n_evts_period = values_main.event.size

        # preparing bins for main axis
        x_main = []
        for i in np.arange(len(bounds_main)-1):
            if np.isinf(bounds_main[i+1]):
                x_main.append( str(self.sci_notation(bounds_main[i])) + '+ '+units['main'] )
            else:
                x_main.append( str(self.sci_notation(bounds_main[i])) + ' to ' + str(self.sci_notation(bounds_main[i+1])) + ' ' + units['main'] )
                
        # preparing bins for supplementary axis
        x_suppl = []
        for j in np.arange(len(bounds_suppl)-1):
            if np.isinf(bounds_suppl[j+1]):
                x_suppl.append( str(self.sci_notation(bounds_suppl[j], opt_integer=integer_legend)) + '+ '+units['suppl'] )
            else:
                x_suppl.append( str(self.sci_notation(bounds_suppl[j], opt_integer=integer_legend)) + ' to ' + str(self.sci_notation(bounds_suppl[j+1], opt_integer=integer_legend)) + ' ' + units['suppl'] )
        
        # preparing values
        weight_counts, count = {}, 0
        for j, s in enumerate(x_suppl):
            tmp = np.zeros(len(x_main))
            for i, m in enumerate(x_main):
                inds = np.where((bounds_main[i] <= values_main.sel(entity=entity).values) & (values_main.sel(entity=entity).values < bounds_main[i+1]) &\
                                (bounds_suppl[j] <= values_suppl.sel(entity=entity).values) & (values_suppl.sel(entity=entity).values < bounds_suppl[j+1]) )[0]
                tmp[i] = len(inds)
            weight_counts[s] = 100 * tmp / n_evts_period # number of events into %
            count += sum(tmp)
        if count != len(values_main.sel(entity=entity).values):
            raise Exception('Bins dont cover all values, enlarge the bounds.')
        
        # preparing size of subplots before text
        xl = [-0.5, len(x_main)-0.5] # doesnt need additional column
        ax_window.set_xlim( xl )
        higher_xmain_categ = np.max( [np.sum([weight_counts[t][i] for t in x_suppl]) for i in np.arange(len(bounds_main)-1)] )
        #yl = [0, (higher_xmain_categ // 10**int(np.log10(higher_xmain_categ)) + 1) * 10**int(np.log10(higher_xmain_categ))] # bound to upper value: eg 64 --> 70; 114 --> 200; etc. Not optimal but works for now.
        yl = [0, 1.05 * higher_xmain_categ]
        ax_window.set_ylim( yl )
        
        # plot
        bottom = np.zeros(len(x_main))
        for i_suppl, suppl in enumerate(weight_counts.keys()):
            _ = ax_window.bar(x=np.arange(len(x_main)), height=weight_counts[suppl], width=0.7, bottom=bottom, label=suppl, facecolor=list_col_suppl[i_suppl] )
            # adding text for each
            for i, x in enumerate(weight_counts[suppl]):
                val = int(np.round(x, 0))
                if val >= 5:# more than 5%
                    t = ax_window.text(x=i, y=bottom[i]+0.5*x, s=str(val)+'%', fontdict={'size':self.fontsizes['pano_txt']},\
                                       verticalalignment='center', horizontalalignment='center', rotation_mode='anchor', bbox={'facecolor':'white', 'alpha':0, 'edgecolor':'white'}, zorder=100)
                    width, pad, shift_y = 0.3, 4, -0.020*(yl[1]-yl[0])#0.3, 1.25, -0.030*(yl[1]-yl[0])
                    height = ((t.get_size()+2*pad)/72.) / (ax_window.get_window_extent().height/self.fig.dpi)
            bottom += weight_counts[suppl]

        # polishing
        ax_window.set_xlabel( 'Multiplication of probability due to entity (-)', size=self.fontsizes['pano_labels'] )
        ax_window.set_ylabel( 'Allocation of events (%)', size=self.fontsizes['pano_axis_labels'] )
        ax_window.set_xticks( ticks=np.arange(len(x_main)), labels=x_main, size=self.fontsizes['pano_axis_ticks'], rotation=-15 )
        #ax_window.set_yticks( ticks=ax_window.get_yticks(), labels=['' for s in ax_window.get_yticks()], size=self.fontsizes['pano_axis_ticks'] )
        if entity in dict_even_shorter_entities:
            lbl = dict_even_shorter_entities[entity]
        elif entity in dict_shorten_entities:
            lbl = dict_shorten_entities[entity]
        else:
            lbl = entity
        ax_window.set_title( '(' + self.dico_panel_entity[entity] + '): ' + lbl, size=self.fontsizes['pano_title'], loc='center', x=0.5, color=self.local_dico_col_entities[entity] )#, fontweight='bold'
        handles, labels = ax_window.get_legend_handles_labels()
        if do_legend:
            legend = ax_window.legend(handles[::-1], labels[::-1], frameon=True, markerfirst=False, loc='center', bbox_to_anchor=(0.5,-0.6), ncols=len(labels), handletextpad=0.1,\
                                      title="Change in intensity due to entity", prop={'size':self.fontsizes['pano_legend']}, title_fontsize=self.fontsizes['pano_legend_title'])
        return ax_window
    #--------------------------------------------------------------
#---------------------------------------------------------------------------------------------------------------------------
#---------------------------------------------------------------------------------------------------------------------------













    





#---------------------------------------------------------------------------------------------------------------------------
# TABLE WITH ALL RESULTS
#---------------------------------------------------------------------------------------------------------------------------
class create_table_all:
    def __init__( self, emdat, pano, emissions_FF, results ):
        # initialization
        self.emdat = emdat
        self.pano = pano
        self.emissions_FF = emissions_FF
        self.results = results

        # preparation
        self.sort_axis_events()
        self.sort_axis_entities()

    #--------------------------------------------------------------
    # PREPARATION
    def sort_axis_events(self):
        # create dico of countries where events occured
        self.dico_countries_events = {}
        for ind_evt in self.emdat.index.values:
            cou = str(self.emdat['Country'].sel(index=ind_evt).values)
            if cou in dico_add_emdatcountries_regionmask:
                cou = dico_add_emdatcountries_regionmask[cou][0]
            if cou not in self.dico_countries_events:
                self.dico_countries_events[cou] = []
            self.dico_countries_events[cou].append( str(self.emdat['DisNo.'].sel(index=ind_evt).values) )
        self.dico_worldregions_events = {}
        for cou in self.dico_countries_events.keys():
            if cou in dico_add_emdatcountries_regionmask:
                cou = dico_add_emdatcountries_regionmask[cou][0]
            reg = dico_country2reg[cou]
            if reg not in self.dico_worldregions_events:
                self.dico_worldregions_events[reg] = []
            self.dico_worldregions_events[reg].append(cou)

        # sorting in each case
        for reg in self.dico_worldregions_events:
            tmp1 = self.dico_worldregions_events[reg]
            tmp1.sort()
            self.dico_worldregions_events[reg] = tmp1
        for cou in self.dico_countries_events:
            tmp2 = self.dico_countries_events[cou]
            tmp2.sort()
            self.dico_countries_events[cou] = tmp2
            
        # create order for regions, countries, events
        self.list_worldregions_events = list(self.dico_worldregions_events.keys())
        self.list_worldregions_events.sort()
        self.list_countries_events, self.list_events = [], []
        for reg in self.list_worldregions_events:
            self.list_countries_events += self.dico_worldregions_events[reg]
            for cou in self.dico_worldregions_events[reg]:
                self.list_events += self.dico_countries_events[cou]

    def sort_axis_entities(self):
        # create dico of countries where carbon majors are headquartered
        self.dico_countries_entities = {}
        for ent in self.emissions_FF.entity.values:
            cou = str(self.emissions_FF['country_entity'].sel(entity=ent).values)
            if cou not in self.dico_countries_entities:
                self.dico_countries_entities[cou] = []
            self.dico_countries_entities[cou].append( ent )
        self.dico_worldregions_entities = {}
        for cou in self.dico_countries_entities.keys():
            if cou in dico_add_emdatcountries_regionmask:
                cou = dico_add_emdatcountries_regionmask[cou][0]
            reg = dico_country2reg[cou]
            if reg not in self.dico_worldregions_entities:
                self.dico_worldregions_entities[reg] = []
            self.dico_worldregions_entities[reg].append(cou)
        
        # sorting in each case
        for reg in self.dico_worldregions_entities:
            tmp1 = self.dico_worldregions_entities[reg]
            tmp1.sort()
            self.dico_worldregions_entities[reg] = tmp1
        for cou in self.dico_countries_entities:
            tmp2 = self.dico_countries_entities[cou]
            tmp2.sort()
            self.dico_countries_entities[cou] = tmp2
            
        # create order for regions, countries, entities
        self.list_worldregions_entities = list(self.dico_worldregions_entities.keys())
        self.list_worldregions_entities.sort()
        self.list_countries_entities, self.list_entities = [], []
        for reg in self.list_worldregions_entities:
            self.list_countries_entities += self.dico_worldregions_entities[reg]
            for cou in self.dico_worldregions_entities[reg]:
                self.list_entities += self.dico_countries_entities[cou]
    #--------------------------------------------------------------


    #--------------------------------------------------------------
    # PREPARATION
    def auto_merge_cells( self, index, row_or_column, pos_start=1, pos_end=np.inf ):
        if row_or_column == 'row':
            icol, do_merge, icol_start = pos_start, False, 1
            while icol <= np.min([pos_end, self.sheet.max_column]):
                if self.sheet.cell(row=index, column=icol).value != '':
                    if do_merge:
                        # merging from previous start to former position
                        self.sheet.merge_cells(start_row=index, end_row=index, start_column=icol_start, end_column=icol-1)
                        # preparing next merge from this position
                        do_merge = False
                    icol_start = icol
                else:
                    do_merge = True
                icol += 1
                    
        elif row_or_column == 'column':
            irow, do_merge, irow_start = pos_start, False, 1
            while irow <= np.min([pos_end, self.sheet.max_row]):
                if self.sheet.cell(row=irow, column=index).value != '':
                    if do_merge:
                        # merging from previous start to former position
                        self.sheet.merge_cells(start_row=irow_start, end_row=irow-1, start_column=index, end_column=index)
                        # preparing next merge from this position
                        do_merge = False
                    irow_start = irow
                else:
                    do_merge = True
                irow += 1

    def prepare_location(self, evt):
        loc = self.results[evt][0].location_init
        if loc in ['', np.nan, 'nan']:
            loc = ''
        return loc
    
    def prepare_period(self, evt):
        dico_tmp = {'Start':[], 'End':[]}
        for sten in ['Start', 'End']:
            for tm in ['Year', 'Month', 'Day']:
                val = str(self.results[evt][0].event_period[sten][tm])
                if ('time_'+sten in results[evt][0].warnings) and (tm in results[evt][0].warnings['time_'+sten]):
                    pass# dont include
                else:
                    if len(val)==1:
                        val = '0'+val
                    dico_tmp[sten].append(val)
        #period_start = self.results[evt][0].format_date(dict_date=self.results[evt][0].event_period['Start'])
        #period_end = self.results[evt][0].format_date(dict_date=self.results[evt][0].event_period['End'])
        #return str(period_start).replace('-','.') + ' - ' + str(period_end).replace('-','.')
        return ('.'.join(dico_tmp['Start'])) + ' - ' + ('.'.join(dico_tmp['End']))
        
    @staticmethod
    def prepare_val(val, type_data):
        if type_data in ['I', 'dI']:
            val_out = str(np.round(val.values, 4)) + ' \u00B0C'
            
        elif type_data in ['PR', 'dPR']:
            limit_PR = 1.e4
            if type_data == 'PR':
                val_tmp = val.values
            else:
                val_tmp = 1+val.values
            if np.isnan(val_tmp) or np.isinf(val_tmp) or val > limit_PR:
                val_out = ">10'000"
            else:
                val_out = str(np.round(val_tmp, 2))
        return val_out
    #--------------------------------------------------------------


    #--------------------------------------------------------------
    # RESULTS
    def write_introduction(self):
        text = [
            "Supplementary table for the scientific publication: Systematic attribution of historical heatwaves to the emissions of carbon majors",\
            "Authors: Yann Quilcaille, Lukas Gudmundsson, Dominik Schumacher, Thomas Gasser, Rick Heede, Quentin Lejeune, Shruti Nath, Wim Thiery, Carl-Friedrich Schleussner, Sonia I. Seneviratne",\
            "Contact: Yann Quilcaille (yann.quilcaille@env.ethz.ch)",\
            " ",\
            "Description: For each one of the 187 heatwaves reported to the disaster database EM-DAT over 2000-2022, an extreme event attribution is performed, to obtain how much anthropogenic climate change has changed the intensity ('Change in intensity') and multiplied the probability ('Probability Ratio') of the heatwave with reference to preindustrial levels. Each extreme event attribution is extended to the emissions of 122 carbon majors, showing how much each carbon major has individually contributed to the intensity ('Contribution to intensity') and multiplied the probability ('Multiplication of preindustrial probability') of the heatwave with reference to preindustrial levels.",\
            "Definition: Extreme Event Attribution: Field of study in climate science that uses physics and statistics to assess the influence of anthropogenic climate change on extreme events, such as heatwaves. More details can be found here: https://www.worldweatherattribution.org/ & https://doi.org/10.5194/ascmo-6-177-2020",\
            "Definition: Carbon Major: Businesses who directly profit from fossil fuel production or other high emitting activities. These entities may be investor-owned, state-owned or nation-states producers. More details can be found here: https://doi.org/10.1007/s10584-013-0986-y",\
            "Disclaimer: The results provided in this table and in the scientific publication are to be considered as scientific information, and not as the quantification of legal responsibilities of actors. This work was driven by two objectives: systematizing extreme event attribution to more events, and extending the attribution to the emitters. The carbon majors were chosen as a category of emitters, because this perspective was less studied by the scientific literature. The entirety of the value chain (Scope 1, 2 and 3) are allocated to the carbon majors, because of data availability.",\
            " ", " "]
        for te in text:
            self.sheet.append( [te] )
        for row in range(1, self.sheet.max_row+1):
            for col in range(1, self.sheet.max_column+1):
                cell = self.sheet.cell(row, col)
                cell.font = openpyxl.styles.Font(bold=True)
    
    def write_header(self):
        # creating header, first rows being: regions, countries, events, confid&median
        row1_tmp = ['Region of disaster', '', '', '']
        row2_tmp = ['Country of disaster', '', '', '']
        row3_tmp = ['Disaster number in EM-DAT', '', '', '']
        row4_tmp = ['Reported location', '', '', '']
        row5_tmp = ['Reported period', '', '', '']
        row6_tmp = ['Best estimate & confidence interval', '', '', '']
        for reg in tmp_tab.list_worldregions_events:
            for cou in tmp_tab.dico_worldregions_events[reg]:
                for evt in tmp_tab.dico_countries_events[cou]:
                    for stat in tmp_tab.range_stats:
                        if (cou == tmp_tab.dico_worldregions_events[reg][0]) and (evt == tmp_tab.dico_countries_events[cou][0]) and (stat == tmp_tab.range_stats[0]):
                            row1_tmp.append(reg)
                        else:
                            row1_tmp.append('')
                        if (evt == tmp_tab.dico_countries_events[cou][0]) and (stat == tmp_tab.range_stats[0]):
                            row2_tmp.append(cou)
                        else:
                            row2_tmp.append('')
                        if (stat == tmp_tab.range_stats[0]):
                            row3_tmp.append(evt)
                            row4_tmp.append( self.prepare_location(evt) )
                            row5_tmp.append( self.prepare_period(evt) )
                        else:
                            row3_tmp.append('')
                            row4_tmp.append('')
                            row5_tmp.append('')
                        row6_tmp.append( stat )
        self.sheet.append( row1_tmp )
        self.sheet.append( row2_tmp )
        self.sheet.append( row3_tmp )
        self.sheet.append( row4_tmp )
        self.sheet.append( row5_tmp )
        self.sheet.append( row6_tmp )
        for row in range(self.sheet.max_row-6+1, self.sheet.max_row+1):
            for col in range(1, self.sheet.max_column+1):
                cell = self.sheet.cell(row, col)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                
        # merging header
        for i in range(1,6+1):
            self.auto_merge_cells( index=self.sheet.max_row-6+i, row_or_column='row')


    def write_attrib_CC(self):
        rowI_tmp = ['Anthropogenic climate change', '', '', 'Change in intensity (\u00B0C)']
        rowPR_tmp = ['', '', '', 'Probability Ratio (-)']
        for evt in self.list_events:
            rowI_tmp.append(self.prepare_val(self.pano.contribs_all['values_global_I_median'].sel(event=evt), type_data='I') )
            rowI_tmp.append(self.prepare_val(self.pano.contribs_all['values_global_I_confid_bottom'].sel(event=evt), type_data='I') + self.junction_CI +\
                            self.prepare_val(self.pano.contribs_all['values_global_I_confid_upper'].sel(event=evt), type_data='I'))
            rowPR_tmp.append(self.prepare_val(self.pano.contribs_all['values_global_PR_median'].sel(event=evt), type_data='PR') )
            rowPR_tmp.append(self.prepare_val(self.pano.contribs_all['values_global_PR_confid_bottom'].sel(event=evt), type_data='PR') + self.junction_CI +\
                             self.prepare_val(self.pano.contribs_all['values_global_PR_confid_upper'].sel(event=evt), type_data='PR'))
        self.sheet.append( rowI_tmp )
        self.sheet.append( rowPR_tmp )
        self.sheet.merge_cells(start_row=self.sheet.max_row-1, end_row=self.sheet.max_row, start_column=1, end_column=3)
    
        # centering everything
        for row in [self.sheet.max_row, self.sheet.max_row-1]:
            for col in range(1, self.sheet.max_column+1):
                cell = self.sheet.cell(row, col)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                if col <= 4:
                    cell.font = openpyxl.styles.Font(bold=True)

    
    def write_buffer_row(self):
        row_tmp = ['Region of headquarter', 'Country of headquarter', 'Carbon major', '']

        # adding empty ones
        for evt in self.list_events:
            for stat in self.range_stats:
                row_tmp.append('')
        self.sheet.append( row_tmp )

        # adding some more bold text
        for col in range(1, 3+1):
            cell = self.sheet.cell(self.sheet.max_row, col)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            cell.font = openpyxl.styles.Font(bold=True)

        # preparing
        self.row_start_merge_cols = self.sheet.max_row

    
    def write_attrib_entity(self, region, country, entity, met_calc):
        # prepare header
        rowI_tmp = []
        if (country == self.dico_worldregions_entities[region][0]) and entity == self.dico_countries_entities[country][0]:
            rowI_tmp.append(region)
        else:
            rowI_tmp.append('')
        if entity == self.dico_countries_entities[country][0]:
            rowI_tmp.append(country)
        else:
            rowI_tmp.append('')
        rowI_tmp.append(entity)
        rowI_tmp.append('Contribution to intensity (\u00B0C)')
        rowPR_tmp = ['', '', '', 'Multiplication of preindustrial probability (-)']

        # fill in
        for evt in self.list_events:
            rowI_tmp.append(self.prepare_val(self.pano.contribs_all['values_'+met_calc+'_I_median'].sel(event=evt, entity=entity), type_data='dI') )
            rowI_tmp.append(self.prepare_val(self.pano.contribs_all['values_'+met_calc+'_I_confid_bottom'].sel(event=evt, entity=entity), type_data='dI') + self.junction_CI +\
                            self.prepare_val(self.pano.contribs_all['values_'+met_calc+'_I_confid_upper'].sel(event=evt, entity=entity), type_data='dI'))
            rowPR_tmp.append(self.prepare_val(self.pano.contribs_all['values_'+met_calc+'_PR_median'].sel(event=evt, entity=entity), type_data='dPR') )
            rowPR_tmp.append(self.prepare_val(self.pano.contribs_all['values_'+met_calc+'_PR_confid_bottom'].sel(event=evt, entity=entity), type_data='dPR') + self.junction_CI +\
                             self.prepare_val(self.pano.contribs_all['values_'+met_calc+'_PR_confid_upper'].sel(event=evt, entity=entity), type_data='dPR'))
        self.sheet.append( rowI_tmp )
        self.sheet.append( rowPR_tmp )

        # centering everything
        for row in [self.sheet.max_row, self.sheet.max_row-1]:
            for col in range(1, self.sheet.max_column+1):
                cell = self.sheet.cell(row, col)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                if col in np.arange(1,4+1):
                    cell.font = openpyxl.styles.Font(bold=True)
    #--------------------------------------------------------------
    
    
    #--------------------------------------------------------------
    # CREATION
    def create( self, path_save, method_calc_entity= 'GMT'):
        # getting confidence interval used everywhere
        disno = list(results.keys())[0]
        tmp = str(100 * np.round( results[disno][-1].dico_q_confid['confid_upper'] - results[disno][-1].dico_q_confid['confid_bottom'], 5))+'%' # avoiding numerical errors
        self.range_stats = ['median', 'CI '+tmp]
        self.junction_CI = ' - '

        # Create a new excel workbook
        self.workbook = openpyxl.Workbook()
        # Select the default sheet
        self.sheet = self.workbook.active

        # create introduction
        self.write_introduction()
        
        # create header
        self.write_header()

        # add results from attribution to climate change
        self.write_attrib_CC()

        # buffer row introducing the header for the carbon majors
        self.write_buffer_row()

        # add results from attribution to entities
        for reg in self.list_worldregions_entities:
            for cou in self.dico_worldregions_entities[reg]:
                for ent in self.dico_countries_entities[cou]:
                    print( ent+' ('+str(self.list_entities.index(ent)+1)+'/'+str(len(self.list_entities))+')'+24*' ', end='\r' )# adding some extra characters because of different length of names.
                    self.write_attrib_entity(reg, cou, ent, method_calc_entity)

        # preparing cleaning
        column_letters = list(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(self.sheet.max_column))
        
        # merging rows that must be & best fit on first columns
        for icol, column_letter in enumerate(column_letters):
            if icol+1 < 4:
                self.auto_merge_cells(index=icol+1, row_or_column='column', pos_start=self.row_start_merge_cols)
            self.sheet.column_dimensions[column_letter].bestFit = True
            
        # saving
        self.workbook.save(os.path.join(path_save, 'table_results.xlsx'))#method_calc_entity
    #--------------------------------------------------------------
#---------------------------------------------------------------------------------------------------------------------------
#---------------------------------------------------------------------------------------------------------------------------







#---------------------------------------------------------------------------------------------------------------------------
# TABLE THAT DETAILS WHICH OBSs & ESMs HAVE BEEN USED FOR EACH EVENT
#---------------------------------------------------------------------------------------------------------------------------
class create_table_data_evts:
    def __init__( self, results, data_to_do ):
        # initialization
        self.results = results
        self.data_to_do = data_to_do

        # preparation
        self.list_events = list(self.results.keys())
        self.list_data = list(self.data_to_do.keys())
    #--------------------------------------------------------------


    #--------------------------------------------------------------
    # RESULTS
    def write_header(self):
        row_tmp = ['Disaster number in EM-DAT'] + self.list_events
        self.sheet.append( row_tmp )
        for col in range(1, self.sheet.max_column+1):
            cell = self.sheet.cell(1, col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    def write_data(self, dat):
        row_tmp = [dat]

        for evt in self.list_events:
            if dat in self.results[evt][4].datasets_postselection:
                row_tmp.append('X')
            else:
                row_tmp.append('')
        self.sheet.append( row_tmp )

        # centering everything
        for col in range(1, self.sheet.max_column+1):
            cell = self.sheet.cell(self.sheet.max_row, col)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            if col == 1:
                cell.font = openpyxl.styles.Font(bold=True)
    #--------------------------------------------------------------
    
    
    #--------------------------------------------------------------
    # CREATION
    def create( self, path_save):
        # Create a new excel workbook
        self.workbook = openpyxl.Workbook()
        # Select the default sheet
        self.sheet = self.workbook.active
        
        # create header
        self.write_header()

        # create one row per event
        for dat in self.list_data:
            self.write_data(dat)

        # preparing cleaning
        column_letters = list(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(self.sheet.max_column))
        for icol, column_letter in enumerate(column_letters):
            self.sheet.column_dimensions[column_letter].bestFit = True
            
        # saving
        self.workbook.save(os.path.join(path_save, 'table_data_events.xlsx'))
    #--------------------------------------------------------------
#---------------------------------------------------------------------------------------------------------------------------
#---------------------------------------------------------------------------------------------------------------------------









